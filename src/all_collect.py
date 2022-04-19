# coding:utf-8
from distutils.archive_util import make_archive
from pickletools import markobject
import unittest
from argparse import Action
from logging import INFO
from re import T
import pandas as pd
from gooey import Gooey, GooeyParser
import sys
import matplotlib
import pandas.plotting._matplotlib
import matplotlib.pyplot as plt
import matplotlib.backends.backend_tkagg
matplotlib.use('tkagg')
# from webdriver_manager.chrome import ChromeDriverManager

    
   	# ......
def Extract_river(data_type='直派水务局',Infilename=None,Outdirname=None):
        # Infilename
        # output_filename=OutDirname
    import codecs
    import re
    import jieba
    jieba.set_dictionary("./data/dict.txt")
    jieba.initialize()
    rDict_filename='./data/river.txt' # 河流名称字典
    stopkey_filename= './data/stopWord_river.txt'# 停用词字典
    tDict_filename='./data/total_dict.txt' # 自建水务词库

    def open_txt(dict):
        '''
        打开文本文档
        '''
        dict=[w.strip() for w in codecs.open(dict, 'r',encoding="utf-8").readlines()] 
        return dict
    
    def txt_read(files):
      txt_dict = {}
      count=0
      fopen = open(files,encoding='utf-8')
      for line in fopen.readlines():
        line = str(line).replace("\n","")
        txt_dict[line.split(' ',1)[0]] = line.split(' ',1)[1]
        #split（）函数用法，逗号前面是以什么来分割，后面是分割成n+1个部分，且以数组形式从0开始
      fopen.close()
      return txt_dict


    def get_content(Infilename):
       '''
       读取excel中指定两列的内容
       '''
       df = pd.read_excel(Infilename)
       content1=[]
       content=[]
        # 按行遍历，读入每行中的两列数据，合并成一个list中一个元素
      # for row in range(2,ws.max_row+1):
    #    for row in range(2,ws.max_row+1):
    #        for col in range(17,18):
    #          content1.append((ws.cell(row,col).value))
    #        content.append(''.join(content1))
    #        content1.clear()
       for i in range(len(df)):
           content1.append(str(df['诉求内容'].iloc[i]))
           content1.append(str(df['主办单位'].iloc[i]))
           
       for j in range(len(content1)):
            if j%2==0:
             content.append(''.join(content1[j:j+2]))
            else:
                continue
       return content
    
    def data_clean():
        '''
        去停用词
        '''
        jieba.load_userdict(tDict_filename) # 添加自建词库
        content=get_content(Infilename)
        stopkey=open_txt(stopkey_filename)
        m={}
        # 按行边历，先分词，再去除无用词
        for i in range(0,len(content)):
          l=[]
          seg = jieba.lcut(content[i])  
          for j in seg:
            if j not in stopkey:  
              l.append(j)
              m[i]=l
        return m

    def search_river():
     '''
     在字典中查找河流名称
     '''
     l = []
     suggestion=data_clean()
     riverdict=txt_read(rDict_filename)
     keys=list(riverdict.keys())
     # 按行遍历，对于每行中的词，与河流词库匹配。
     for j in range(len(suggestion)):
            for i in suggestion[j]:
                count=0
                for k in riverdict.keys():
                #  if i in riverdict[k]:
                    c=re.search(i,riverdict[k])
                    a=list(riverdict[k])
                    if c==None or len(i)<2:
                      continue
                    elif c.re.pattern==k :
                       l.append(k)
                       count+=1
                       break
                    elif  a[min(c.span())-1]!=" "  :
                       continue
                    elif a[max(c.span())]==' ':
                       l.append(k)
                       count+=1
                       break  
                if(count==1):
                 break              
            if(count==1):
               continue
            else:
                l.append('不详')
     return l
        
    # if __name__=='__main__':
    '''
    输出结果到excel
    '''
    df = pd.read_excel(Infilename)
    df['具体点位']=search_river()
    savepath = Outdirname + "\\直派水务局河湖增补.xlsx"  #导入excel数据，设置表头标题
    writer = pd.ExcelWriter(savepath)   #设置导出的excel文件地址
    df.to_excel(writer, sheet_name="Sheet1", index=False)  
    writer.save()
    sys.exit()
def Extract_village(data_type='直派水务局',Infilename=None,Outdirname=None):
    import jieba
    import pandas as pd
    jieba.set_dictionary("./data/dict.txt")
    jieba.initialize()
    import difflib
    class Nstr:
        def __init__(self, arg):
            self.x = arg

        def __sub__(self, other):
            c = self.x.replace(other.x, "")
            return c

    def get_equal_rate(str1, str2):
        return difflib.SequenceMatcher(None, str1, str2).quick_ratio()

    def getpoint(strindex):
        global resualt
        cunming = seg[strindex]
        scorelist1 = []
        scorelist2 = []
        scorelist3 = []
        scorelist4 = []
        for cun in item_seg:
            scorelist1.append(get_equal_rate(cunming, cun))
        # print(max(scorelist1))
        # print(len(cunming))
        # print(item_seg[scorelist1.index(max(scorelist1))])
        if max(scorelist1) == 1 and len(cunming) != 1:
            cunmingindex = scorelist1.index(max(scorelist1))
            cunming1 = item_seg[cunmingindex]
            resualt = cunming1
        else:
            cunming = seg[strindex - 1] + seg[strindex]
            for cun in item_seg:
                scorelist2.append(get_equal_rate(cunming, cun))
            # print(max(scorelist2))
            if max(scorelist2) > 0.9:
                cunmingindex = scorelist2.index(max(scorelist2))
                cunming2 = item_seg[cunmingindex]
                resualt = cunming2
            else:
                cunming = seg[strindex - 2] + seg[strindex - 1] + seg[strindex]
                for cun in item_seg:
                    scorelist3.append(get_equal_rate(cunming, cun))
                # print(max(scorelist3))
                if max(scorelist3) > 0.8:
                    cunmingindex = scorelist3.index(max(scorelist3))
                    cunming3 = item_seg[cunmingindex]
                    resualt = cunming3
                else:
                    cunming = seg[strindex - 3] + seg[strindex - 2] + seg[strindex - 1] + seg[strindex]
                    for cun in item_seg:
                        scorelist4.append(get_equal_rate(cunming, cun))
                    # print(max(scorelist4))
                    if max(scorelist4) > 0.75:
                        cunmingindex = scorelist4.index(max(scorelist4))
                        cunming4 = item_seg[cunmingindex]
                        resualt = cunming4
                    else:
                        resualt = "未找到"
        return resualt

    atxt = r"./data/北京市全部地名.txt"
    btxt = r'./data/北京市街道.txt'
    ctxt = r"./data/北京市城区.txt"
    zhuizhongres = []
    qulist = []
    listzhen = []
    countlist = []
    xiangzhen = []
    mainkey = ["村", "小区", "胡同", "社区", "里", "路", "园", "院", "楼", "区"]
    keywordlibiao = ["村", "小区", "胡同", "社区", "里", "路", "园", "院", "楼", "区"]
    global seg
    global item_seg
    jieba.load_userdict(atxt)
    item_seg = []
    with open(atxt, encoding='utf-8') as f:
        for line in f:
            item_seg.append(line.strip('\n'))
    with open(btxt, encoding='utf-8') as f:
        for line in f:
            line = line.strip('\n')
            xiangzhen.append(line)
    with open(ctxt, encoding='utf-8') as cleardata:
        for qu in cleardata:
            qulist.append(qu.strip('\n'))
    item_seg = list(set(item_seg) - set(xiangzhen) - set(qulist))
    if data_type == "直派行业":
        file_name = Infilename
        data = pd.read_excel(file_name, sheet_name='Sheet1')
        jiedao = data["被反映街乡镇"]
        dianwei = data["问题点位"]
        neirong = data["主要内容"]
        for i in range(0, len(data)):
            tiquresult = "不详"
            keywordlist = []
            reslist = []
            chun = dianwei[i]
            chengqu = jiedao[i]
            nr = neirong[i]
            seg_list = jieba.cut(nr)
            seg = list(seg_list)
            for ct in range(0, 3):
                chun = Nstr(chun) - Nstr("北京市")
                chun = Nstr(chun) - Nstr(chengqu)
                for item in xiangzhen:
                    chun = Nstr(chun) - Nstr(item)
            reschun = chun
            for key in mainkey:
                r = chun.find(key)
                if r == -1:
                    pass
                else:
                    tiquresult = chun[0:r + len(key)]
                    for biaozhunhua in item_seg:
                        if get_equal_rate(biaozhunhua, tiquresult) > 0.8:
                            tiquresult = biaozhunhua
                            break
                        else:
                            pass
                    break
            if tiquresult == "不详":
                for keyword in keywordlibiao:
                    for segitem in seg:
                        if keyword in segitem:
                            keywordlist.append(seg.index(segitem))
                countkeyword = len(keywordlist)
                if countkeyword == 0:
                    tiquresult = "不详"
                else:
                    for keyworditem in range(0, countkeyword):
                        res = getpoint(keywordlist[keyworditem])
                        reslist.append(res)
                    panduan = 0
                    for resu in reslist:
                        for keyword in keywordlibiao:
                            if keyword in resu:
                                tiquresult = resu
                                panduan = 1
                                break
                            else:
                                pass
                        if panduan == 1:
                            break
                    if panduan == 0:
                        tiquresult = reschun
                        for biaozhunhua in item_seg:
                            if get_equal_rate(biaozhunhua, tiquresult) > 0.8:
                                tiquresult = biaozhunhua
                                break
                            else:
                                pass
            zhuizhongres.append(tiquresult)
        data['小区村庄'] = zhuizhongres
        savepath = Outdirname + "\\直派行业小区村庄增补.xlsx"
        writer = pd.ExcelWriter(savepath)
        data.to_excel(writer, sheet_name="Sheet1", index=False)
        writer.save()
    elif data_type == "直派水务局":
        file_name = Infilename
        data = pd.read_excel(file_name)
        neirong = data["诉求内容"]
        for i in range(0, len(data)):
            tiquresult = "不详"
            keywordlist = []
            reslist = []
            nr = neirong[i]
            seg_list = jieba.cut(nr)
            seg = list(seg_list)
            for keyword in keywordlibiao:
                for segitem in seg:
                    if keyword in segitem:
                        keywordlist.append(seg.index(segitem))
            countkeyword = len(keywordlist)
            if countkeyword == 0:
                tiquresult = "不详"
            else:
                for keyworditem in range(0, countkeyword):
                    res = getpoint(keywordlist[keyworditem])
                    reslist.append(res)
                panduan = 0
                for resu in reslist:
                    for keyword in keywordlibiao:
                        if keyword in resu:
                            tiquresult = resu
                            panduan = 1
                            break
                        else:
                            pass
                    if panduan == 1:
                        break
                if panduan == 0:
                    tiquresult = "不详"
                    for biaozhunhua in item_seg:
                        if get_equal_rate(biaozhunhua, tiquresult) > 0.8:
                            tiquresult = biaozhunhua
                            break
                        else:
                            pass
            zhuizhongres.append(tiquresult)
        data['小区村庄'] = zhuizhongres
        savepath = Outdirname + "\\直派水务局小区村庄增补.xlsx"
        writer = pd.ExcelWriter(savepath)
        data.to_excel(writer, sheet_name="Sheet1", index=False)
        writer.save()
def Extract_attribute(Infilename=None,Outdirname=None):
    import difflib

    import jieba
    jieba.set_dictionary("./data/dict.txt")
    jieba.initialize()
    def get_equal_rate_1(str1, str2):
        return difflib.SequenceMatcher(None, str1, str2).quick_ratio()

    def stop_word(textlist, stopwordlist):
        outstrlist = []
        for word in textlist:
            if word not in stopwordlist:
                if word != "\t":
                    outstrlist.append(word)
        return outstrlist

    df = pd.read_excel(Infilename, sheet_name='区街道乡镇')
    cklist=['市政供水','自备井','物业','自来水公司','小区','村']
    for i in cklist:
        jieba.add_word(i)

    def stopwordslist():
        stopwords=['$', '0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '?', '_', '“', '”', '、', '。', '《', '》', '一', '一些', '一何', '一切', '一则', '一方面', '版权', '那', '一样', '一般', '一转眼', '万一', '上', '上下', '下', '不', '偏心', '自己', '不光', '不单', '不只', '不外乎', '不如', '面临', '不尽', '不尽然', '不得', '不怕', '不惟', '不成', '不拘', '不料', '是', '不比', '吃的', '不特', '不独', '不屑一顾', '不至于', '不若', '或者', '不过', '不问', '与', '与其', '与其说', '与否', '本体', '且', '且不说', '且说', '重新', '个', '购买', '临', '为', '为了', '为什么', '为何', '直到', '其中', '为着', '乃', '喜爱', '喜爱于', '么', '之', '一个', '休息', '之类', '乌乎', '乎', '乘', '也', '也好', '也罢', '了', '二来', '于', '于是', '于是乎', '云云', '云尔', '些', '亦', '人', '像这样的', '家人', '什么', '不好看', '今', '附近', '仍', '依旧', '从', '回顾', '从而', '他', '生者', '她们', '以', '以上', '以为', '以便', '向往', '以及', '以故', '以期', '以来', '以至', '以至于', '以致', '我们', '任', '任何', '任凭', '似的', '但', '但凡', '但是', '何', '何以', '何况', '在哪里', '什么时候', '余外', '成为', '你', '你们', '使', '过去', '例如', '依', '相比', '根据', '方便', '俺', '俺们', '以后', '就使', '以后或', '然然', '如果若', '借', '假使', '镜像', '假若', '傥然', '像', '儿', '先不先', '光是', '简体中文', '全部', '兮兮', '关于', '其', '其一', '有', '其二', '其他', '其余', '其他', '即将', '一', '具体说来', '兼之', '内', '再', '再其次', '再则', '再有', '再者', '再者说', '那些', ',', '冲', '况且', '几', '几时', '凡', '凡是', '凭直觉', '综合', '用心', '出来', '分别', '则', '则甚', '别', '别人', '别处', '别是', '别的', '别管', '别说', '到', '相对', '前此', '前者', '加之', '许可', '即', '即令', '即使', '轻微的', '即如', '即或', '即若', '却', '去', '又', '又及', '及', '以及', '及至', '反之', '可能性', '回复', '回复说', '受到', '另', '同时', '另外', '另悉', '只', '只当', '只怕', '只是', '只有', '只消', '给', '只限', '叫', '叮咚', '可', '可以', '而', '可见', '各', '每个人', '各位', '各种', '自己', '同', '同时', '后', '后期', '向', '向使', '向着', '吓', '吗', '否则', '吧', '吧哒', '吱', '呀', '呃', '呕吐', '呗', '呜呜', '呜呼', '呢', '呵', '呵呵', '呸', '呼哧', '咋', '和', '咚', '咦', '那', '咱', '我们', '咳', '哇', '哈', '哈哈', '哉', '哎', '哎呀', '哎哟', '哗', '哟', '哦哦', '哩', '哪', '哪个', '哪些', '哪儿', '天哪', '哪年', '相对', '哪样', '哪边', '哪里', '哼', '哼唷', '唉', '唯有', '啊', '啐', '啥', '啦', '啪达', '啷当', '喂', '喏', '喔唷', '喽', '嗡', '嗡嗡', '嗬', '嗯', '嗳', '嘎', '嘎登', '嘘', '嘛嘛', '嘻', '嘿', '嘿嘿', '因', '因为', '因了', '因此', '因着', '减少', '固然', '在', '在下', '在于', '地', '因为', '报价', '多', '怎么样', '多少', '大', '大家', '她', '女', '好', '如', '如上', '那个', '如下', '怎么办', '如其', '王', '如是', '如果如果', '如此', '如若', '始而', '哪个料', '哪个知', '宁', '宁可', '避免', '宁肯', '它', '的', '对', '对于', '对待', '对方', '对比', '将', '小', '尔', '尔后', '尔尔', '尚且', '就', '就是', '就是了', '就是说', '不良', '她', '尽', '虽然', '……', '岂但', '己', '已', '已矣', '巴', '巴巴', '并', '并且', '不是', '庶乎', '庶几', '开外', '开始', '归归', '归齐', '当', '当地', '当然', '当着', '彼', '彼时', '彼此', '往', '待', '很', '得', '得了', '怎么', '怎么', '怎么办', '怎么样', '怎奈', '怎么样', '花朵', '总的来看', '总', '总的说来', '总说之', '恰好相反', '您的', '惟其', '慢说', '我', '我们', '或', '或则', '求你了', '或曰', '或者', '达到', '所', '所以', '属于', '所幸', '所有', '才', '不', '打', '打从', '把', '抑或', '拿', '按', '按照', '水', '想', '据', '集中', '外', '故', '故此', '故而', '旁人', '无', '无宁', '图', '既', '既往', '既是', '既然', '时候', '是', '是', '是的', '曾', '替', '替补', '最', '有', '今人', '有关', '有及', '有时', '有的', '望', '朝', '想', '本', '我', '本地', '本着', '做作', '来', '来着', '来自', '说', '极了', '果然', '果真', '某', '某人', '部分', '某某', '可知', '欤', '正值', '如', '正巧', '正是', '此', '此地', '这里', '另外', '此时', '来源', '此间', '免宁', '每', '颗粒', '比', '比及', '前辈', '比方', '没奈何', '沿', '沿线', '漫说', '焉', '然则', '然后', '然而', '照', '照着', '犹且', '犹自', '甚且', '什么', '甚或', '甚而', '甚至', '甚至于', '用', '解释', '由', '因为', '由是', '本人', '点', '的', '确实', '话', '直到', '相对说', '省得', '看', '眨眼', '着', '着呢', '矣', '矣乎', '矣哉', '离', '竟而', '第', '等', '候诊室', '等等', '简言之', '管', '类如', '紧接着', '纵', '纵令', '纵使', '纵然', '经', '经过', '结果', '给', '继之', '继后', '继而', '综上所述', '罢了', '者', '而', '还有', '而况', '而后', '而外', '笑声', '混合', '说的', '能', '能不能', '腾', '自', '自个儿', '因为', '自各儿', '自后', '自家', '自己', '自打', '自我', '至', '至于', '来自', '至若', '致', '般的', '若', '若夫', '允', '若果', '若非', '莫吃', '莫如', '莫若', '虽', '虽则', '是不是', '虽说', '被', '要', '要不', '要不是', '要吃', '是否', '可以', '暗示', '不愿如', '让', '从', '论', '设使', '设或', '设若', '诚如', '诚然', '该', '说来', '诸', '诸位', '日常', '谁', '谁人', '谁料', '谁知', '贼死', '赖以', '赶', '起', '起见', '趁', '趁着', '越是', '距', '跟', '较', '较之', '边', '过', '还', '还是', '还有', '还有', '这', '这来', '这个', '如此', '这么些', '这么样', '这么点儿', '这', '这会儿', '这里', '这就是说', '那时', '这样', '这次', '这般', '这篇', '这里', '过', '连', '地图', '逐步', '通过', '遵循', '遵照', '那', '那个', '那么', '那些', '那样', '那些', '那会儿', '那边', '那时', '那', '那般', '反', '那里', '都', '其他人', '开始', '针对', '阿', '除', '除了', '除外', '除开', '梦想', '通过', '随', '迎接', '随时', '随着', '难道说', '非但', '非徒', '非特', '非独', '靠', '顺', '顺着', '首先', '！', '，', '：', '；', '？']
        return stopwords
    list1= []
    list1.extend(0 for _ in range(len(df)))
    list2 = []
    list2.extend(0 for _ in range(len(df)))

    for i in range(len(df)):

        cell_context = df['主要内容'].iloc[i]
        cut_content = jieba.lcut(cell_context)

        stopwords = stopwordslist()  # 创建停用词列表
        fin_cut_content = stop_word(cut_content, stopwords)


        list3 = []
        for s in fin_cut_content:

            if get_equal_rate_1("市政供水", s) > 0.8 or get_equal_rate_1("物业", s) > 0.9 or get_equal_rate_1("自来水公司",
                                                                                        s) > 0.8 or get_equal_rate_1(
                    "小区", s) == 1 or get_equal_rate_1("自来水集团",s) > 0.8:
                list3.append("市政供水")
            if get_equal_rate_1("物业", s) > 0.9 :
                list3.append("物业")
            if get_equal_rate_1("自来水公司", s) > 0.8:
                list3.append("自来水公司")
            if get_equal_rate_1("自来水集团", s) > 0.8:
                list3.append("自来水公司")
            if get_equal_rate_1("自备井", s) == 1 :
                list3.append(s)
            if get_equal_rate_1("井",s) == 1 or get_equal_rate_1("井水",s) > 0.9 or get_equal_rate_1("乡村",s) > 0.9 or get_equal_rate_1("村民",s) > 0.9:
                list3.append("自备井")

        if "自备井" in list3 :

            list1[i] = "自备井"
            list2[i] = "其他产权单位"
        elif"市政供水" in list3 or "物业" in list3 or "自来水公司" in list3:
            list1[i] = "市政供水"
            if "物业" in list3:
                list2[i] = "物业"
            elif "自来水公司" in list3:

                list2[i] = "自来水公司"
            else:
                list2[i] = "其他产权单位"
        else :
            list1[i] = "未知或其他产权单位"
            list2[i] = "未知或其他产权单位"
        #     list2.append("不详")

    df['供水类型'] = list1
    df['负责单位'] = list2
    savepath=Outdirname+"\\供水属性标注.xlsx"
    writer = pd.ExcelWriter(savepath)
    df.to_excel(writer, sheet_name="区街道乡镇", index=False)
    writer.save()
    sys.exit()
def Sum_Extract_time(data_type='直派水务局',Infilename=None,Outdirname=None):
    import pandas as pd
    if data_type=='直派水务局':
        zs=pd.read_excel(Infilename)
        zsdatadf=zs['登记时间']
        zs['年']=zsdatadf.str[2:4]
        zs['月']=zsdatadf.str[5:7]
        zs['日']=zsdatadf.str[8:10]
        zs['时']=zsdatadf.str[11:13]
        old_zs=pd.read_excel('./data/直派水务局（总）.xlsx')
        all_zs=old_zs.append(zs,ignore_index=True)
        savepath1='./data/直派水务局（总）.xlsx'
        writer1 = pd.ExcelWriter(savepath1)   #设置导出的excel文件地址
        writer3 = pd.ExcelWriter(Outdirname+'\\直派水务局（总）.xlsx')
        all_zs.to_excel(writer1,index=False)
        all_zs.to_excel(writer3,index=False)
        writer1.save()
        writer3.save()
    elif data_type=='直派行业':
        hy=pd.read_excel(Infilename)
        hydatadf=hy['创建时间']
        hy['年']=hydatadf.str[2:4]
        hy['月']=hydatadf.str[5:7]
        hy['日']=hydatadf.str[8:10]
        hy['时']=hydatadf.str[11:13]
        old_hy=pd.read_excel('./data/直派行业（总）.xlsx')
        all_hy=old_hy.append(hy,ignore_index=True)
        savepath2='./data/直派行业（总）.xlsx'
        writer2 = pd.ExcelWriter(savepath2)   #设置导出的excel文件地址
        writer4 = pd.ExcelWriter(Outdirname+'\\直派行业（总）.xlsx')
        all_hy.to_excel(writer2,index=False)
        all_hy.to_excel(writer4,index=False)
        writer2.save()
        writer4.save()
def Classification(Infilename=None,Outdirname=None):
    pass

def Extract_reason(Infilename=None,Outdirname=None):
    import difflib

    import jieba
    jieba.set_dictionary("./data/dict.txt")
    jieba.initialize()
    import pandas as pd

    def get_equal_rate(str1, str2):
        return difflib.SequenceMatcher(None, str1, str2).quick_ratio()

    class Nstr:
        def __init__(self, arg):
            self.x = arg

        def __sub__(self, other):
            c = self.x.replace(other.x, "")
            return c

    def frequency_sort(items):
        # your code here
        lst1 = []
        for i in items:
            if i not in lst1:
                lst1.append(i)
        lst = []
        dic = {}
        # # print(lst1)
        for i in lst1:
            dic[i] = items.count(i)
        dic = dict(sorted(dic.items(), key=lambda x: x[1], reverse=True))
        # #print(dic)
        count = 0
        for k, v in dic.items():
            for i in range(v):
                lst.append(k)
        # # print(lst)
        return lst
    file_name = Infilename
    data = pd.read_excel(file_name, sheet_name='Sheet1')
    jieguo = data["处理结果"]
    mainkey = ["因", "因为", "由", "由于", "原因"]
    listkeyword = []
    resendlist = []
    reasonlist = []
    for i in range(0, len(data)):
        prores = jieguo[i]
        if str(prores) == "None":
            resend = "无数据"
        else:
            seg_list = jieba.cut(prores)
            seg = list(seg_list)
            # print(seg)
            for keyword in mainkey:
                for item in seg:
                    if item == keyword:
                        listkeyword.append(item)
            listkeyword = frequency_sort(listkeyword)
            if len(listkeyword) != 0:
                proresnew = prores
                for keyw in listkeyword:
                    keywordcount = listkeyword.count(keyw)
                    if keywordcount != 1:
                        findk = proresnew.find(keyw)
                        finde = proresnew.find("。", findk)
                        resend = proresnew[findk:finde + 1]
                        resendlist.append(resend)
                        proresnew = proresnew[finde + 1:]
                    else:
                        findkk = prores.find(keyw)
                        findee = prores.find("。", findkk)
                        resendd = prores[findkk:findee + 1]
                        resendlist.append(resendd)
                strresend = ""
                for i in resendlist:
                    strresend = strresend + i
                resend = strresend
            else:
                resend = "未找到"
        reasonlist.append(resend)
        resendlist = []
        listkeyword = []
        # print(listkeyword)
    data['诉求原因'] = reasonlist
    savepath = Outdirname + "\\诉求原因提取.xlsx"
    writer = pd.ExcelWriter(savepath)
    data.to_excel(writer, sheet_name="Sheet1", index=False)
    writer.save()

def Count(data_type='直派水务局',YearMonth=None,count_type=None,InDirname=None,Outdirname=None):
    import operator
    import os
    from collections import Counter

    import openpyxl
    import pandas as pd
    from openpyxl import load_workbook

    filepath1 = 'demo_excel/办公单位.xlsx'
    filepath2 = 'demo_excel/诉求类型一级表.xlsx'
    # filepath3 = './tongji_out/统计报告.xlsx'
    # filepath4 = './tongji_out/诉求风险趋势变化表.xlsx'
    filepath3 = 'demo_excel/out_demo/统计报告.xlsx'
    filepath4 = 'demo_excel/out_demo/诉求风险趋势变化表.xlsx'
    filepath5 = 'demo_excel/街道二级诉求.xlsx'
    filepath6 = 'demo_excel/各区街道集团统计样板.xlsx'
    filepath7 = Outdirname + '\\各区街道集团统计.xlsx'
    filepath8 = Outdirname + '\\统计报告.xlsx'
    filepath9 = Outdirname + '\\诉求风险趋势变化表.xlsx'
    year=int(YearMonth[:-2])
    month=int(YearMonth[-2:])
    # filepath7 = Outdirname+'./town_out/各区街道集团统计.xlsx'
    # filepath8 = './demo_excel/汇总样板.xlsx'    #createxcel 的filepath3、4
    # filepath9 = './demo_excel/诉求类型样板.xlsx'

    filelist1 = ['2021.1','2021.2','2021.3','2021.4','2021.5','2021.6','2021.7','2021.8','2021.9','2021.10','2021.11','2021.12','2022.1','2022.2','2022.3','2022.4','2022.5','2022.6','2022.7','2022.8','2022.9','2022.10','2022.11','2022.12','2023.1','2023.2','2023.3','2023.4','2023.5','2023.6','2023.7','2023.8','2023.9','2023.10','2023.11','2023.12']
    if year==2020:
        filelist1=filelist1[:month]
    elif year==2021:
        filelist1=filelist1[:12+month]
    elif year==2022:
        filelist1=filelist1[:24+month]
    elif year==2023:
        filelist1=filelist1[:36+month]
    excel_num1 = len(filelist1)
    # print(filelist, excel_num,filelist[excel_num-3].replace('.xlsx','') )  # 这里filelist从0开始，到（excel数量-1）结束

    def getexcel_town(choice, excel_filepath, dict_filepath):
        wb = openpyxl.Workbook()  # 创建一个新的excel
        we = wb['Sheet']
        if choice == '各区街道集团':
            we.title = '统计'
            we['A1'] = '各区街道集团'
            we['A2'] = '二级问题'
        month_num = excel_num1
        for i in range(0, month_num):
            we.cell(2, 2 + i).value = filelist1[i]
            we.cell(2, month_num + 2).value = '合计'
            i += 1
        for j in range(0, len(getdict(dict_filepath))):
            we.cell(3 + j, 1).value = getdict(dict_filepath)[j]['名称']
            j += 1

        first_num = j + 4
        if choice == '各区街道集团':
            we.cell(first_num, 1).value = '三级问题（排名降序）'  #
        # second_num=first_num+j+3                   #  已解决+剔除+挂账 部分
        wb.save(excel_filepath)  # 保存表格
        return (first_num)

    def alltongji_town(choice, excel_start, first_path, second_path, third_path):  #### 程序运行时不要打开用到的 Excel ！

        rb = openpyxl.load_workbook(first_path)
        if choice == '各区街道集团':
            wl = rb['统计']
        for i in range(excel_start - 1, excel_num1):  # 决定统计表格起始    *********************************************
            if choice == '各区街道集团':
                new_num = getexcel_town('各区街道集团', first_path, second_path)
            # data = pd.read_excel('./data_town/' + filelist1[i], sheet_name='各区街道集团')
            data = pd.read_excel(InDirname, sheet_name='各区街道集团')
            row = len(data)  # 表格行数
            x_pandas_list = data['二级问题']  # 专业情况
            c = Counter(x_pandas_list)
            # print(c)
            list_len = len(c)
            factor_len = len(getdict(second_path))  # 办理单位数量
            # print(factor_len)
            print(f"{choice}{filelist1[i]}开始统计···")
            all_tongji, not_tongji, other_tongji, type_all_tongji, type_other_tongji = 0, 0, 0, 0, 0
            all_list = {}
            not_list = {}
            other_list = {}
            type_all_list = {}
            type_other_list = {}  # 存储每月各单位的各类型数量
            kong_num = 0  # 办公单位为空
            for j in range(0, factor_len):  # 循环所有办公单位
                all_num = 0  # 全部类型  非权属  已解决+剔除+挂账 （等）
                all_data = []
                for k in range(0, list_len):
                    exit_flag = 0
                    if choice == '各区街道集团':
                        if format(c.most_common()[k][0].replace('（', '').replace('）', '')) == format(
                                getdict(second_path)[j]['名称'].replace('(', '').replace(')',
                                                                                       '')):  # 二级问题所在字典  一个去中文（） ,一个去英文()
                            exit_flag = 1
                            if exit_flag == 1:
                                all_num = c.most_common()[k][1]
                            if exit_flag == 0:
                                all_num = 0
                ##  结果存入数组
                all_list[j] = all_num
                all_data.append(all_list)
                all_tongji += all_num
                # print(all_num, all_tongji, row ,j)
            for j in range(0, factor_len):
                if choice == '各区街道集团':
                    #   数量
                    wl.cell(3 + j, 2 + i).value = int_str(all_data[0][j])  # 统计结果写入 汇总表
            #   保存
            rb.save(third_path)

            if all_tongji != row:
                if choice == '各区街道集团':
                    print(
                        f"此月份表格有{row - all_tongji}条二级问题信息未添加到‘街道二级诉求.xlsx’文件中，或有的名字打错！！！")  # 这里格式要求严格，如果名字没错也没有空的，那么注意名字前后的 ‘ ’，即，空格！
            print(f"{filelist1[i]}统计完成\n")

        dic = []
        for j in range(0, factor_len):
            sum_number = 0
            for k in range(excel_start - 1, excel_num1):
                if choice == '各区街道集团':
                    sum_number += int(wl.cell(3 + j, 2 + k).value)  # 统计结果写入 汇总表
            wl.cell(3 + j, 2 + excel_num1).value = sum_number
            dic.append({'name': getdict(second_path)[j]['名称'], 'number': sum_number})
        rb.save(third_path)
        sorted_x = sorted(dic, key=operator.itemgetter('number'), reverse=True)
        print(f"{sorted_x} ,{len(sorted_x)}\n")

        #### 二级问题排序，创建动态表格框架
        first_num = new_num + 1
        for top_num in range(0, len(sorted_x)):
            wl['A' + int_str(new_num + 1)] = f"第{top_num + 1}位:" + sorted_x[top_num]['name']
            if sorted_x[top_num]['name'] == '中水':
                zs_num = new_num + 3
            if sorted_x[top_num]['name'] == '供水':
                gs_num = new_num + 7
            for top_i in range(0, excel_num1):
                wl.cell(new_num + 1, 2 + top_i).value = filelist1[top_i].replace('.xlsx', '')
                top_i += 1
            town_getdict_len = len(town_getdict(sorted_x[top_num]['name']))  # mou 某二级问题问题的三级字典长度
            for top_j in range(0, town_getdict_len):
                wl.cell(new_num + 2 + top_j, 1).value = town_getdict(sorted_x[top_num]['name'])[top_j]['名称']
                top_j += 1
            new_num += top_j + 2
        # 保存
        rb.save(third_path)
        ##################################################################################
        #### 开始前十类写入
        for top_num in range(0, len(sorted_x)):
            for i in range(excel_start - 1, excel_num1):
                # data = pd.read_excel('./data_town/' + filelist1[i], sheet_name='各区街道集团')
                data = pd.read_excel(InDirname, sheet_name='区街道乡镇')
                # row = len(data)  # 表格行数
                x_pandas_list = data['三级问题']  # 专业情况
                c = Counter(x_pandas_list)
                list_len = len(c)
                factor_len = len(town_getdict(sorted_x[top_num]['name']))  # 三级问题数量
                print(f"第{top_num + 1}位{sorted_x[top_num]['name']}{filelist1[i]}开始统计···")
                all_tongji = 0
                all_list = {}
                for j in range(0, factor_len):  # 循环所有办公单位
                    all_num, not_num, other_num = 0, 0, 0  # 全部类型  非权属  已解决+剔除+挂账 （等）
                    all_data = []
                    for k in range(0, list_len):
                        exit_flag = 0
                        if choice == '各区街道集团':
                            if c.most_common()[k][0] == town_getdict(sorted_x[top_num]['name'])[j]['名称']:
                                exit_flag = 1
                                if exit_flag == 1:
                                    all_num = c.most_common()[k][1]
                                if exit_flag == 0:
                                    all_num = 0
                        ##  结果存入数组
                    all_list[j] = all_num
                    all_data.append(all_list)
                    all_tongji += all_num
                    # print(all_num, all_tongji, row ,j)
                for j in range(0, factor_len):
                    if choice == '各区街道集团':
                        #   数量
                        wl.cell(first_num + 1 + j, 2 + i).value = int_str(all_data[0][j])  # 统计结果写入 汇总表
                ###  找到具有共同三级的二级问题的类的所在  行
                for find_i in range(0, len(getdict(second_path))):
                    if wl.cell(row=find_i + 3, column=1).value == sorted_x[top_num]['name']:
                        the_row = find_i + 3
                        # print(the_row,excel_start,wl.cell(row=the_row,column=i+2 ).value,type(wl.cell(row=the_row,column= i+2).value),all_tongji )
                ####  修改冗余部分
                if all_tongji > int(wl.cell(row=the_row, column=i + 2).value):
                    if sorted_x[top_num]['name'] == '中水':
                        error_num = all_tongji - int(
                            wl.cell(row=the_row, column=i + 2).value)  # 因有共同三级问题情况多出来的   目前只知道缴费
                        if choice == '各区街道集团':
                            wl.cell(zs_num, 2 + i).value = int_str(all_data[0][1] - error_num)  # 统计结果写入 汇总表
                            all_tongji = all_tongji - error_num
                            # print(the_row,wl.cell(row=the_row,column=i+2 ).value,all_data [0][1],all_tongji,error_num)
                    if sorted_x[top_num]['name'] == '供水':
                        error_num = all_tongji - int(
                            wl.cell(row=the_row, column=i + 2).value)  # 因有共同三级问题情况多出来的   目前只知道缴费
                        if choice == '各区街道集团':
                            wl.cell(gs_num, 2 + i).value = int_str(all_data[0][5] - error_num)  # 统计结果写入 汇总表
                            all_tongji = all_tongji - error_num
                            # print(the_row,wl.cell(row=the_row,column=i+2 ).value,all_data[0][5],all_tongji ,error_num)
                #   保存
                rb.save(third_path)

                if all_tongji == int(wl.cell(row=the_row, column=i + 2).value):
                    if choice == '各区街道集团':
                        print(f"此月份表格有{all_tongji}条{sorted_x[top_num]['name']}三级问题信息。")
                #  再次判断是否有共同三级问题情况
                if all_tongji > int(wl.cell(row=the_row, column=i + 2).value):
                    if choice == '各区街道集团':
                        print(
                            f"此月份表格有{all_tongji - int(wl.cell(row=the_row, column=i + 2).value)}条三级问题信息在其他二级问题的下属三级问题中也存在！！")
                if all_tongji < int(wl.cell(row=the_row, column=i + 2).value):  # sorted_x[top_num]['number']:
                    if choice == '各区街道集团':
                        print(f"此月份表格有{int(wl.cell(row=the_row, column=i + 2).value) - all_tongji}条三级问题信息未被统计！！！")
                print(f"{sorted_x[top_num]['name']}{filelist1[i]}统计完成\n")
            first_num += 2 + factor_len
        print('统计完成！')

    def getexcel(choice, excel_filepath, dict_filepath):
        # list_month = ['2020.1', '2020.2', '2020.3', '2020.4', '2020.5', '2020.6', '2020.7', '2020.8', '2020.9',
        #               '2020.10', '2020.11', '2020.12', '2021.1', '2021.2', '2021.3', '2021.4', '2021.5', '2021.6',
        #               '2021.7', '2021.8', '2021.9', '2021.10', '2021.11', '2021.12']

        wb = openpyxl.Workbook()  # 创建一个新的excel
        we = wb['Sheet']
        we2 = wb.create_sheet('sheet2', 1)
        we3 = wb.create_sheet('sheet3', 2)

        if choice == '单位':
            
            we3.title = '水利工程运行维护'
            we2.title = '占比'
            we.title = '数量'
            we['A1'] = '全部诉求类型'
            we2['A1'] = '全部诉求类型'
            we3['A1'] = '道路破损'
            we['A2'] = '单位'
            we2['A2'] = '单位'
            we3['A2'] = '单位'
        elif choice == '诉求':
            we2.title = '占比'
            we.title = '数量'
            we['A1'] = '全部诉求'
            we['A2'] = '诉求类型（一级诉求）'
            we2['A1'] = '全部诉求'
            we2['A2'] = '诉求类型（一级诉求）'
        month_num = excel_num
        for i in range(0, month_num):
            we.cell(2, 2 + i).value = filelist[i]
            we2.cell(2, 2 + i).value = filelist[i]
            we3.cell(2, 2 + i).value = filelist[i]
            i += 1
        for j in range(0, len(getdict(dict_filepath))):
            we.cell(3 + j, 1).value = getdict(dict_filepath)[j]['名称']
            we2.cell(3 + j, 1).value = getdict(dict_filepath)[j]['名称']
            we3.cell(3 + j, 1).value = getdict(dict_filepath)[j]['名称']
            j += 1
        first_num = j + 4
        if choice == '单位':
            we.cell(first_num, 1).value = '非权属类型'  # 非权属部分   相邻部分空一格
            we.cell(first_num + 1, 1).value = '单位'
            we2.cell(first_num, 1).value = '非权属类型'  # 非权属部分   相邻部分空一格
            we2.cell(first_num + 1, 1).value = '单位'
            we3.cell(first_num, 1).value = '路灯'  # 路灯
            we3.cell(first_num + 1, 1).value = '单位'
        elif choice == '诉求':
            we.cell(first_num, 1).value = '非权属诉求'  # 非权属部分   相邻部分空一格
            we.cell(first_num + 1, 1).value = '诉求类型（一级诉求）'
            we2.cell(first_num, 1).value = '非权属诉求'  # 非权属部分   相邻部分空一格
            we2.cell(first_num + 1, 1).value = '诉求类型（一级诉求）'
        for i in range(0, month_num):
            we.cell(first_num + 1, 2 + i).value = filelist[i]
            we2.cell(first_num + 1, 2 + i).value = filelist[i]
            we3.cell(first_num + 1, 2 + i).value = filelist[i]
            i += 1
        for j in range(0, len(getdict(dict_filepath))):
            we.cell(2 + first_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            we2.cell(2 + first_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            we3.cell(2 + first_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            j += 1
        second_num = first_num + j + 3  # 已解决+剔除+挂账 部分
        if choice == '单位':
            we.cell(second_num, 1).value = '已解决+剔除+挂账'  # 非权属部分
            we.cell(second_num + 1, 1).value = '单位'
            we2.cell(second_num, 1).value = '已解决+剔除+挂账'  # 非权属部分
            we2.cell(second_num + 1, 1).value = '单位'
            we3.cell(second_num, 1).value = '路边乱停车'  # 非权属部分
            we3.cell(second_num + 1, 1).value = '单位'
        elif choice == '诉求':
            we.cell(second_num, 1).value = '已解决+剔除+挂账'  # 非权属部分
            we.cell(second_num + 1, 1).value = '诉求类型（一级诉求）'
            we2.cell(second_num, 1).value = '已解决+剔除+挂账'  # 非权属部分
            we2.cell(second_num + 1, 1).value = '诉求类型（一级诉求）'
        for i in range(0, month_num):
            we.cell(second_num + 1, 2 + i).value = filelist[i]
            we2.cell(second_num + 1, 2 + i).value =filelist[i]
            we3.cell(second_num + 1, 2 + i).value = filelist[i]
            i += 1
        for j in range(0, len(getdict(dict_filepath))):
            we.cell(2 + second_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            we2.cell(2 + second_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            we3.cell(2 + second_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            j += 1

        if choice == '单位':
            third_num = second_num + j + 3
            we.cell(third_num, 1).value = '水环境维护类 总诉求'  # 水环境维护类 总诉求
            we.cell(third_num + 1, 1).value = '单位'
            we2.cell(third_num, 1).value = '水环境维护类 总诉求'  # 水环境维护类 总诉求
            we2.cell(third_num + 1, 1).value = '单位'
            for i in range(0, month_num):
                we.cell(third_num + 1, 2 + i).value = filelist[i]
                we2.cell(third_num + 1, 2 + i).value = filelist[i]
                i += 1
            for j in range(0, len(getdict(dict_filepath))):
                we.cell(2 + third_num + j, 1).value = getdict(dict_filepath)[j]['名称']
                we2.cell(2 + third_num + j, 1).value = getdict(dict_filepath)[j]['名称']
                j += 1
            fourth_num = third_num + j + 3
            we.cell(fourth_num, 1).value = '水环境维护类 已解决+剔除+挂账'  # 水环境维护类 已解决+剔除+挂账
            we.cell(fourth_num + 1, 1).value = '单位'
            we2.cell(fourth_num, 1).value = '水环境维护类 已解决+剔除+挂账'  # 水环境维护类 已解决+剔除+挂账
            we2.cell(fourth_num + 1, 1).value = '单位'
            for i in range(0, month_num):
                we.cell(fourth_num + 1, 2 + i).value = filelist[i].replace('.xlsx', '')
                we2.cell(fourth_num + 1, 2 + i).value = filelist[i].replace('.xlsx', '')
                i += 1
            for j in range(0, len(getdict(dict_filepath))):
                we.cell(2 + fourth_num + j, 1).value = getdict(dict_filepath)[j]['名称']
                we2.cell(2 + fourth_num + j, 1).value = getdict(dict_filepath)[j]['名称']
                j += 1
        wb.save(excel_filepath)  # 保存表格
        return (first_num)

    def getdict(filepath):
        book = load_workbook(filepath)
        sheet = book.active
        rows = sheet.rows
        headers = [cell.value for cell in next(rows)]
        all_rows = []
        for row in rows:
            # 创建一个空字典来存放这里面的值
            data = {}
            for title, cell in zip(headers, row):
                data[title] = cell.value
            all_rows.append(data)
        # print(len(all_rows),all_rows[62]['名称'])
        # print(all_rows )
        return (all_rows)

    def town_getdict(type):
        book = openpyxl.load_workbook(f"./demo_excel/街道三级诉求/{type}三级诉求.xlsx")
        # sheet = book.active
        sheet = book['Sheet1']
        rows = sheet.rows
        headers = [cell.value for cell in next(rows)]
        all_rows = []
        for row in rows:
            # 创建一个空字典来存放这里面的值
            data = {}
            for title, cell in zip(headers, row):
                data[title] = cell.value
            all_rows.append(data)
        # print(len(all_rows),all_rows[62]['名称'])
        # print(all_rows   )
        return (all_rows)

    def int_str(int_num):
        lst = []
        s = ''
        if int_num == 0:
            s = '0'
        while int_num > 0:
            num = int_num % 10
            lst.append(num)
            int_num //= 10
        for i in lst[::-1]:
            s += str(i)
        # print(s)
        return s

    def percent(num1, num2):
        a = '{:.2%}'.format(num1 / num2)
        # print(a)
        return (a)
    # monthlist=['']
    filelist = ['2020.1','2020.2','2020.3','2020.4','2020.5','2020.6','2020.7','2020.8','2020.9','2020.10','2020.11','2020.12','2021.1','2021.2','2021.3','2021.4','2021.5','2021.6','2021.7','2021.8','2021.9','2021.10','2021.11','2021.12','2022.1','2022.2','2022.3','2022.4','2022.5','2022.6','2022.7','2022.8','2022.9','2022.10','2022.11','2022.12','2023.1','2023.2','2023.3','2023.4','2023.5','2023.6','2023.7','2023.8','2023.9','2023.10','2023.11','2023.12']
    if year==2020:
        filelist=filelist[:month]
    elif year==2021:
        filelist=filelist[:12+month]
    elif year==2022:
        filelist=filelist[:24+month]
    elif year==2023:
        filelist=filelist[:36+month]
    # for root, dirs, files in os.walk("./data", topdown=False):
    # for root, dirs, files in os.walk(InDirname, topdown=False):
    #     for name in files:
    #         str2 = os.path.join(name)  # str = os.path.join(root, name)

    #         if str2.split('.')[-1] == 'xlsx':
    #             filelist.append(str2)
    excel_num = len(filelist)

    # print(filelist, excel_num)  # 这里filelist从0开始，到（excel数量-1）结束

    def alltongji(choice, first_path, second_path, third_path):  #### 程序运行时不要打开用到的 Excel ！
        rb = openpyxl.load_workbook(first_path)
        if choice == '水利工程运行维护' or choice=='单位':
            wl = rb['数量']
            wl2 = rb['占比']
            wl3 = rb['水利工程运行维护']
        if choice == '诉求':
            wl = rb['数量']
            wl2 = rb['占比']
        for i in range(excel_num - 1, excel_num):  # 决定统计表格起始    *********************************************
            if choice == '单位':
                new_num = getexcel('单位', first_path, second_path)
            elif choice == '诉求':
                new_num = getexcel('诉求', first_path, second_path)
            elif choice == '水利工程运行维护':
                new_num = getexcel('单位', first_path, second_path)  # 因为同样是统计个办公单位的
            factor_column = 0  ##  '主办单位'所在列
            result_column = 0  ##
            data = openpyxl.load_workbook(InDirname)
            sheets = data.sheetnames
            sheet = data[sheets[0]]
            # sheet = data["sheet1"]
            column = len(sheet[1])  # 表格列数
            row = len(tuple(sheet))  # 表格行数
            for column_num in range(1, column + 1):
                if sheet.cell(row=1, column=column_num).value == '主办单位':
                    # print(sheet[1][column_num-1].value)
                    factor_column = column_num
            # print(sheet.cell(row=12,column=factor_column).value)     #测试空格单元 类型是 ‘ ‘ 还是None，none会出错。
            for column_num in range(1, column + 1):
                if sheet.cell(row=1, column=column_num).value == '解决类型':
                    # print(sheet[1][column_num-1].value)
                    result_column = column_num
            # print(sheet.cell(row=41, column=result_column).value)
            for column_num in range(1, column + 1):
                if sheet.cell(row=1, column=column_num).value == '诉求类型一级':
                    # print(sheet[1][column_num-1].value)
                    type_column = column_num
            for column_num in range(1, column + 1):
                if sheet.cell(row=1, column=column_num).value == '诉求类型三级':
                    third_type_column = column_num
            # print( factor_column, result_column,type_column,row,column)
            factor_len = len(getdict(second_path))  # 办理单位数量
            # print(factor_len)
            print(f"{choice}{filelist[i]}开始统计···")
            all_tongji, not_tongji, other_tongji, type_all_tongji, type_other_tongji = 0, 0, 0, 0, 0
            all_list = {}
            not_list = {}
            other_list = {}
            type_all_list = {}
            type_other_list = {}  # 存储每月各单位的各类型数量
            kong_num = 0  # 办公单位为空
            for j in range(0, factor_len):  # 循环所有办公单位
                all_num = 0  # 全部类型
                not_num = 0  # 非权属
                other_num = 0  # 已解决+剔除+挂账 （等）
                type_all_num, type_other_num, type_not_num, kong_num1, kong_num2 = 0, 0, 0, 0, 0
                all_data = []
                not_data = []
                other_data = []
                type_all_data = []
                type_other_data = []
                for k in range(2, row + 1):
                    if choice == '单位':
                        if format(sheet.cell(row=k, column=factor_column).value.replace('（', '').replace('）',
                                                                                                         '')) == format(
                                getdict(second_path)[j]['名称'].replace('(', '').replace(')',
                                                                                       '')):  # 办理单位所在字典  一个去中文（） ,一个去英文()
                            all_num += 1
                            if format(sheet.cell(row=k, column=result_column).value.replace('（', '').replace('）',
                                                                                                             '')) == '非权属':
                                not_num += 1
                            elif format(sheet.cell(row=k, column=result_column).value.replace('（', '').replace('）',
                                                                                                               '')) == '已解决' or '剔除' or '挂账' or '' or '主动剔除' or 'none' or '正在办理':
                                other_num += 1
                            if sheet.cell(row=k, column=type_column).value == '水环境维护':
                                type_all_num += 1
                                if format(sheet.cell(row=k, column=result_column).value.replace('（', '').replace('）',
                                                                                                                 '')) == '非权属':
                                    type_not_num += 1
                        elif format(
                                sheet.cell(row=k, column=factor_column).value.replace('（', '').replace('）', '')) == '':
                            kong_num1 += 1
                        elif sheet.cell(row=k, column=factor_column).value == 'none':
                            kong_num2 += 1
                        type_other_num = type_all_num - type_not_num

                    elif choice == '诉求':
                        if format(sheet.cell(row=k, column=type_column).value.replace('（', '').replace('）',
                                                                                                       '')) == format(
                                getdict(second_path)[j]['名称'].replace('(', '').replace(')',
                                                                                       '')):  # 办理单位所在字典  一个去中文（） ,一个去英文()
                            all_num += 1
                            if format(sheet.cell(row=k, column=result_column).value.replace('（', '').replace('）',
                                                                                                             '')) == '非权属':
                                not_num += 1
                            elif format(sheet.cell(row=k, column=result_column).value.replace('（', '').replace('）',
                                                                                                               '')) == '已解决' or '剔除' or '挂账' or '' or '主动剔除' or 'none' or '正在办理':
                                other_num += 1
                        elif format(
                                sheet.cell(row=k, column=type_column).value.replace('（', '').replace('）', '')) == '':
                            kong_num1 += 1
                        elif sheet.cell(row=k, column=type_column).value == 'none':
                            kong_num2 += 1
                        type_other_num = type_all_num - type_not_num

                    elif choice == '水利工程运行维护':
                        if format(sheet.cell(row=k, column=factor_column).value.replace('（', '').replace('）',
                                                                                                         '')) == format(
                                getdict(second_path)[j]['名称'].replace('(', '').replace(')',
                                                                                       '')):  # 办理单位所在字典  一个去中文（） ,一个去英文()
                            if sheet.cell(row=k, column=type_column).value == '水利工程运行维护':
                                if sheet.cell(row=k, column=third_type_column).value == '道路破损':
                                    all_num += 1
                                if sheet.cell(row=k, column=third_type_column).value == '路灯':
                                    not_num += 1
                                if sheet.cell(row=k, column=third_type_column).value == '路边乱停车':
                                    other_num += 1
                        elif format(
                                sheet.cell(row=k, column=factor_column).value.replace('（', '').replace('）', '')) == '':
                            kong_num1 += 1
                        elif sheet.cell(row=k, column=factor_column).value == 'none':
                            kong_num2 += 1
                        type_other_num = type_all_num - type_not_num

                ##  结果存入数组
                all_list[j] = all_num
                all_data.append(all_list)
                not_list[j] = not_num
                not_data.append(not_list)
                other_list[j] = other_num
                other_data.append(other_list)
                type_all_list[j] = type_all_num
                type_all_data.append(type_all_list)
                type_other_list[j] = type_other_num
                type_other_data.append(type_other_list)
                all_tongji += all_num
                not_tongji += not_num
                other_tongji += other_num
                type_all_tongji += type_all_num
                type_other_tongji += type_other_num
                kong_num = kong_num1 + kong_num2
                print(all_num, not_num, other_num, all_tongji, not_tongji, other_tongji, kong_num, type_all_num,
                      type_other_num, row - 1, j)
                if type_all_tongji == 0:
                    type_all_tongji = 1
                if type_other_tongji == 0:
                    type_other_tongji = 1

            for j in range(0, factor_len):
                if choice == '单位':
                    month_num = excel_num
                    for ii in range(0, month_num):
                        wl.cell(2, 2 + ii).value = filelist[ii]
                        wl2.cell(2, 2 + ii).value = filelist[ii]
                        wl3.cell(2, 2 + ii).value = filelist[ii]

                        wl.cell(new_num + 1, 2 + ii).value = filelist[ii]
                        wl2.cell(new_num + 1, 2 + ii).value = filelist[ii]
                        wl3.cell(new_num + 1, 2 + ii).value = filelist[ii]

                        wl.cell(2 * new_num  , 2 + ii).value = filelist[ii]
                        wl2.cell(2 * new_num , 2 + ii).value = filelist[ii]
                        wl3.cell(2 * new_num , 2 + ii).value = filelist[ii]

                        ######单位独有
                        wl.cell(3 * new_num- 1, 2 + ii).value = filelist[ii]
                        wl2.cell(3 * new_num- 1, 2 + ii).value = filelist[ii]

                        wl.cell(4 * new_num - 2, 2 + ii).value = filelist[ii]
                        wl2.cell(4 * new_num - 2, 2 + ii).value = filelist[ii]
                    #   数量
                    wl.cell(3 + j, 2 + i).value = int_str(all_data[0][j])  # 统计结果写入 汇总表
                    wl.cell(new_num + 2 + j, 2 + i).value = int_str(not_data[0][j])
                    wl.cell(2 * new_num + 1 + j, 2 + i).value = int_str(other_data[0][j])
                    #   占比
                    wl2.cell(3 + j, 2 + i).value = percent(all_data[0][j], all_tongji)  # 统计结果写入 汇总表
                    wl2.cell(new_num + 2 + j, 2 + i).value = percent(not_data[0][j], not_tongji)
                    wl2.cell(2 * new_num + 1 + j, 2 + i).value = percent(other_data[0][j], other_tongji)
                    #   水环境维护
                    wl.cell(3 * new_num + j, 2 + i).value = int_str(type_all_data[0][j])
                    wl.cell(4 * new_num - 1 + j, 2 + i).value = int_str(type_other_data[0][j])
                    wl2.cell(3 * new_num + j, 2 + i).value = percent(type_all_data[0][j], type_all_tongji)
                    wl2.cell(4 * new_num - 1 + j, 2 + i).value = percent(type_other_data[0][j], type_other_tongji)

                elif choice == '诉求':
                    month_num = excel_num
                    for ii in range(0, month_num):
                        wl.cell(2, 2 + ii).value = filelist[ii]
                        wl2.cell(2, 2 + ii).value = filelist[ii]


                        wl.cell(new_num + 1, 2 + ii).value = filelist[ii]
                        wl2.cell(new_num + 1, 2 + ii).value = filelist[ii]


                        wl.cell(2 * new_num, 2 + ii).value = filelist[ii]
                        wl2.cell(2 * new_num, 2 + ii).value = filelist[ii]


                    #   数量
                    wl.cell(3 + j, 2 + i).value = int_str(all_data[0][j])  # 统计结果写入 汇总表
                    wl.cell(new_num + 2 + j, 2 + i).value = int_str(not_data[0][j])
                    wl.cell(2 * new_num + 1 + j, 2 + i).value = int_str(other_data[0][j])
                    #   占比
                    wl2.cell(3 + j, 2 + i).value = percent(all_data[0][j], all_tongji)  # 统计结果写入 汇总表
                    wl2.cell(new_num + 2 + j, 2 + i).value = percent(not_data[0][j], not_tongji)
                    wl2.cell(2 * new_num + 1 + j, 2 + i).value = percent(other_data[0][j], other_tongji)

                elif choice == '水利工程运行维护':
                    month_num = excel_num
                    for ii in range(0, month_num):
                        wl.cell(2, 2 + ii).value = filelist[ii]
                        wl2.cell(2, 2 + ii).value = filelist[ii]
                        wl3.cell(2, 2 + ii).value = filelist[ii]

                        wl.cell(new_num + 1, 2 + ii).value = filelist[ii]
                        wl2.cell(new_num + 1, 2 + ii).value = filelist[ii]
                        wl3.cell(new_num + 1, 2 + ii).value = filelist[ii]

                        wl.cell(2 * new_num, 2 + ii).value = filelist[ii]
                        wl2.cell(2 * new_num, 2 + ii).value = filelist[ii]
                        wl3.cell(2 * new_num, 2 + ii).value = filelist[ii]

                        ######单位独有
                        wl.cell(3 * new_num - 1, 2 + ii).value = filelist[ii]
                        wl2.cell(3 * new_num - 1, 2 + ii).value = filelist[ii]

                        wl.cell(4 * new_num - 2, 2 + ii).value = filelist[ii]
                        wl2.cell(4 * new_num - 2, 2 + ii).value = filelist[ii]
                    #   数量
                    wl3.cell(3 + j, 2 + i).value = int_str(all_data[0][j])  # 统计结果写入 汇总表
                    wl3.cell(new_num + 2 + j, 2 + i).value = int_str(not_data[0][j])
                    wl3.cell(2 * new_num + 1 + j, 2 + i).value = int_str(other_data[0][j])
            #   保存
            rb.save(third_path)
            if choice == '单位':
                rb.save(filepath3)
            if choice == '诉求':
                rb.save(filepath4)
            if choice == '水利工程运行维护':
                rb.save(filepath3)

            if (all_tongji + kong_num) != row - 1:
                if choice == '单位':
                    print(
                        f"此月份表格有{row - 1 - all_tongji - kong_num}条办公单位信息未添加到‘办公单位.xlsx’文件中，或有的单位名字打错。")  # 这里格式要求严格，如果名字没错也没有空的，那么注意名字前后的 ‘ ’，即，空格！
                elif choice == '诉求':
                    print(f"此月份表格有{row - 1 - all_tongji - kong_num}条一级诉求类型信息未添加到‘诉求类型一级.xlsx’文件中，或有的诉求类型名字打错。")
                elif choice == '水利工程运行维护':
                    print(f"水利工程运行维护--此月份表格有{all_tongji}条关于道路破损，{not_tongji}条关于道路破损，{other_tongji}条关于道路破损。")
            if kong_num != 0:
                if choice == '单位':
                    print(f"{filelist[i]}表中办公单位为空的数据有{kong_num}条,包括’ ‘和’none‘！")
                elif choice == '诉求':
                    print(f"{filelist[i]}表中一级诉求类型为空的数据有{kong_num}条,包括’ ‘和’none‘！")
                elif choice == '水利工程运行维护':
                    print(f"水利工程运行维护--{filelist[i]}表中办公单位为空的数据有{kong_num}条,包括’ ‘和’none‘！")
            print(f"{filelist[i]}统计完成")
        # rb.save(last_filepath)
        print('统计完成！')

    # if __name__ == '__main__':
    if  count_type=='单位解决类型':
        #  用于统计各单位全部诉求类型、非全属、解决等数量和占比，还有水环境维护类的全部诉求类型、非全属、解决等数量和占比。
        alltongji('单位', filepath3, filepath1, filepath8)  # 汇总样板表、办公单位、统计报告
    if count_type == '诉求变化趋势':
        #  用于统计诉求类型（一级诉求）的全部诉求类型、非全属、解决等数量和占比
        alltongji('诉求', filepath4, filepath2, filepath9)  # 诉求类型样板、诉求类型一级表、诉求风险趋势变化表
    if count_type == '水利工程运行维护':
        #  用于统计水利工程运行维护里各单位对道路破损、路灯、路边乱停车的数量统计
        alltongji('水利工程运行维护', filepath3, filepath1, filepath8)  # 汇总样板表、办公单位、统计报告
    if count_type == '各区街道集团':
        getexcel_town('各区街道集团', filepath6, filepath5)  # 样板
        #  用于统计各区街道集团的二级问题诉求，同时对统计的二级问题进行排序，然后对排序后的二级问题对应的三级问题进行统计
        alltongji_town('各区街道集团', 1, filepath6, filepath5, filepath7)  # 标题、样表、起始月份、字典、输出表

def Draw(data_type='直派水务局',draw_type='折线图（诉求变化趋势）',Infilename=None,Outdirname=None):
    # import matplotlib.pyplot as plt
    plt.rcParams['font.sans-serif'] = ['SimHei']
    import os
    if not os.path.exists(Outdirname+'\\绘图\\'):
        os.makedirs(Outdirname+'\\绘图\\')
    def scrdf(df,indexname,judge):#
        df0=df[indexname]
        indexs = []
        for index, value in enumerate(df0):
            if judge == value:
                indexs.append(index)
        scrdf = pd.DataFrame(df.iloc[indexs].values)
        scrdf.columns=df.columns
        return scrdf
    def Draw_line(name):
        def draw_line_base(month,datadf,yearlist,length):
            # plt.figure()
            plt.plot()
            step = int(length/12)
            # print(length,step)
            for i in range(step):
                # print("%d",i*12)
                plt.plot(month[:12],datadf[int(i*12):int((i+1)*12)], marker='o', markersize=3)
                for a, b in zip(month[:12], datadf[int(i*12):int((i+1)*12)]):
                    plt.text(a, b, b, ha='center', va='bottom', fontsize=10)  # 设置数据标签位置及大小
            plt.plot(month[:length-step*12],datadf[step*12:length],marker='o',markersize=3)
            for a, b in zip(month[:length-step*12], datadf[step*12:length]):
                plt.text(a, b, b, ha='center', va='bottom', fontsize=10)
            plt.legend(yearlist[:step+1])  # 设置折线名称
            # plt.show()
        df=pd.read_excel('./data/诉求风险趋势变化表.xlsx')[-32:]
        plt.rcParams['font.sans-serif'] = ['SimHei']
        ptdf = df.iloc[df['全部诉求'].values.tolist().index(name)]
        datadf = ptdf.values.tolist()
        datadf.remove(name)
        datadf = list(map(int, datadf))
        yearlist=['2020年','2021年','2022年','2023年','2024年','2025年','2026年','2027年','2028年']
        month = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']
        length=len(datadf)
        draw_line_base(month,datadf,yearlist,length)
        plt.title(label=name)
        plt.xlabel('时间')  # x轴标题
        plt.ylabel('诉求值')  # y轴标题
        plt.savefig(Outdirname  + '\\绘图\\'+name+'诉求变化趋势图.jpg')
    def Draw_pie(data_type='直派水务局'): # make a square figure
        df=pd.read_excel(Infilename)
        if data_type=='直派水务局':
            pt0=df['诉求类型一级'].value_counts()
            pt1=df['诉求类型二级'].value_counts()
            pt2=df['诉求类型三级'].value_counts()
        elif data_type=='直派行业':
            pt0=df['一级问题'].value_counts()
            pt1=df['二级问题'].value_counts()
            pt2=df['三级问题'].value_counts()
        # For China, make the piece explode a bit
        expl = [0.1,0.05,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]#第二块即China离开圆心0.1
        # Colors used. Recycle if not enough.
        colors  = ["firebrick","indianred","tomato","rosybrown","lightcoral","salmon","sienna","sandybrown",'peru','greenyellow','palegreen','springgreen','aquamarine','cyan','deepskyblue',"lightskyblue",'violet','hotpink',"coral","magenta","green","lightgreen","orange"]  #设置颜色（循环显示）
        # Pie Plot
        # autopct: format of "percent" string;百分数格式
        lb0=pt0.index.tolist()
        qs0=pt0.values.tolist()
        lb1=pt1.index.tolist()
        qs1=pt1.values.tolist()
        lb2=pt2.index.tolist()
        qs2=pt2.values.tolist()
        if len(lb0)>=15:
            qs00=[]
            qs00.extend(qs0[:15])
            qs00.append(sum(qs0[15:]))
            lb00=[]
            lb00.extend(lb0[:15])
            lb00.append('其余诉求')
            plt.figure(1, figsize=(9,6))
            plt.pie(qs00,colors= colors,explode=expl[:len(qs00)], labels=lb00, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局一级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right",bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局一级诉求-类型占比分布图.png')
            plt.close()
        else:
            plt.figure(1, figsize=(9,6))
            plt.pie(qs0,colors= colors,explode=expl[:len(qs0)], labels=lb0, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局一级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right",bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局一级诉求-类型占比分布图.png')
            plt.close()
        if len(lb1)>=20:
            qs01=[]
            qs01.extend(qs1[:20])
            qs01.append(sum(qs1[20:]))
            lb01=[]
            lb01.extend(lb1[:20])
            lb01.append('其余诉求')
            plt.figure(1, figsize=(9,6))
            plt.pie(qs01,colors= colors,explode=expl[:len(qs01)], labels=lb01, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局二级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right", 
                            bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局二级诉求-类型占比分布图.png')
            plt.close()
        else:
            plt.figure(1, figsize=(9,6))
            plt.pie(qs1,colors= colors,explode=expl[:len(qs1)], labels=lb1, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局二级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right",bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局二级诉求-类型占比分布图.png')
            plt.close()
        if len(lb2)>=20:
            qs02=[]
            qs02.extend(qs2[:20])
            qs02.append(sum(qs2[20:]))
            lb02=[]
            lb02.extend(lb2[:20])
            lb02.append('其余诉求')
            plt.figure(1, figsize=(9,6))
            plt.pie(qs02,colors= colors,explode=expl[:len(qs02)], labels=lb02, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局三级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right", 
                            bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局三级诉求-类型占比分布图.png')
            plt.close()
        else:
            plt.figure(1, figsize=(9,6))
            plt.pie(qs2,colors= colors,explode=expl[:len(qs2)], labels=lb2, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局三级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right",bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局三级诉求-类型占比分布图.png')
            plt.close()
    def Draw_bar(data_type='直派水务局'):
        
        def draw_bar(pt,name):
            plt.rcParams['font.sans-serif'] = ['SimHei']
            # plt.figure(14.0,8.0)
            pt.plot.bar()
            x = range(0, len(pt.index.tolist()), 1)
            plt.xticks(x,rotation=60)
            plt.subplots_adjust(bottom=0.35)
            plt.title(name)
            for a, b in zip(x, pt.values.tolist()):
                    plt.text(a, b+0.05,'%d' %b, ha='center', va='bottom', fontsize=10)
            plt.savefig(Outdirname + '\\绘图\\'+name+'-诉求柱状图.jpg')
            plt.close()
        if data_type=='直派水务局':
            df=pd.read_excel(Infilename)
            pt = df['诉求类型一级'].value_counts()
            draw_bar(pt,'直派水务局诉求类型一级')
            pt1list=pt.index.tolist()
            for i in range(len(pt1list)):
                draw_bar(scrdf(df,'诉求类型一级',pt1list[i])['诉求类型二级'].value_counts(),'直派水务局诉求类型二级-'+pt1list[i])
        elif data_type=='直派行业':
            df=pd.read_excel(Infilename)
            pt = df['二级问题'].value_counts()
            draw_bar(pt,'直派行业二级问题')
            pt1list=pt.index.tolist()
            for i in range(len(pt1list)):
                draw_bar(scrdf(df,'二级问题',pt1list[i])['三级问题'].value_counts(),'直派行业三级问题-'+pt1list[i])
    def Draw_geo():
        from pyecharts import options as opts
        from pyecharts.charts import Geo, Map
        from pyecharts.render import make_snapshot
        from snapshot_selenium import snapshot as driver
        def draw_geo(df,name):
            quname=['密云区','延庆区','朝阳区','丰台区','石景山区','海淀区','门头沟区','房山区','通州区','顺义区','昌平区','大兴区','怀柔区','平谷区','东城区','西城区']
            ptindex=df['被反映区'].value_counts().index.tolist()
            # ptindex=list(set(ptindex+quname))
            ptindex=list((ptindex+ [item for item in quname if str(item) not in ptindex]))
            ptvalue=df['被反映区'].value_counts().values.tolist()
            if len(ptindex)!=len(ptvalue):
                ptvalue.extend(0 for _ in range(len(ptindex)-len(ptvalue)))
            maxvalue=max(map(int,ptvalue))
            while (maxvalue %5):
                maxvalue+=1
            c = (
                Map()
                .add("", [list(z) for z in zip(ptindex, ptvalue)], "北京",zoom=1.25,aspect_scale=0.9,layout_center=['60%','60%'],
                label_opts=opts.LabelOpts(is_show=True,position='outsideleft',font_size=10,color='#708069',rotate = '30',horizontal_align = 'center',font_weight='bold',vertical_align ='middle'))
                .set_global_opts(
                title_opts=opts.TitleOpts(title=name),
                visualmap_opts=opts.VisualMapOpts(type_='color', orient='vertical',max_=maxvalue,
                                                    is_piecewise=True,range_color=['#71ae9b','#8bb08e', '#c1d389','#e3d76b','#f7b44a','#f27127','#cc8775']
                ))
            )
            make_snapshot(driver, c.render(), Outdirname +'\\pic\\'+ name+".png")
            b = (
                Map()
                .add("", [list(z) for z in zip(ptindex, ptvalue)], "北京",zoom=1.25,aspect_scale=0.9,layout_center=['60%','60%'],
                label_opts=opts.LabelOpts(is_show=True,position='outsideleft',font_size=10,color='#708069',rotate = '30',horizontal_align = 'center',font_weight='bold',vertical_align ='middle'))
                .set_global_opts(
                title_opts=opts.TitleOpts(title=name),
                visualmap_opts=opts.VisualMapOpts(type_='color', orient='vertical',is_piecewise=True,pieces=[{'min':0,'max':10,'label':'[0-10)',"color":'#71ae9b'},{'min':11,'max':30,'label':'[10-30)',"color":'#8bb08e'},{'min':31,'max':100,'label':'[30-100)',"color":'#c1d389'},{'min':101,'max':200,'label':'[100-200)',"color":'#e3d76b'},{'min':201,'max':350,'label':'[200-350)',"color":'#f7b44a'},
                                                            {'min':350,'label':'[350,)',"color":'#f27127'}]
                                                            # {'min':0,'max'=10,label:'0-10','#cc8775']
                                                    )
            ))
            make_snapshot(driver, b.render(), Outdirname +'\\pic\\'+ name +".png")
        df=pd.read_excel(Infilename)
        pt = df['三级问题'].value_counts()
        pt1list=pt.index.tolist()
        for i in range(10):
            draw_geo(scrdf(df,'三级问题',pt1list[i]),'直派行业三级诉求-'+pt1list[i]+'-地理分布图')# if __name__=='__main__':
    if draw_type=='水利工程运行维护变化趋势-折线图':
        Draw_line('水利工程运行维护')
    elif draw_type=='水环境维护变化趋势-折线图':
        Draw_line('水环境维护')
    elif draw_type=='雨水与海绵城市变化趋势-折线图':
        Draw_line('雨水与海绵城市')
    elif draw_type=='饼状图':
        Draw_pie(data_type)
    elif draw_type=='柱状图（直派水务局一级诉求）' or draw_type=='柱状图（直派行业三级问题）'  :
        Draw_bar(data_type)
    elif draw_type=='地理分布图（区）' and data_type=='直派行业':
            Draw_geo()

def similar_ayalyse(YearMonth='202201',Infilename=None,Outdirname=None):
    import difflib
    data_df=pd.read_excel(Infilename)
    data=data_df['诉求内容']
    excel_shape=data_df.shape[0]
    def string_similar(s1,s2):
        return difflib.SequenceMatcher(None, s1, s2).quick_ratio()
    def similar_count():
        nums=[]
        for i in range(excel_shape):
            j=i+1
            while(j<excel_shape):
                simi_rate=string_similar(data[i],data[j])
                print(simi_rate)
                if(simi_rate>=0.7):
                    nums.append(i)
                    nums.append(j)
                    data[j]='0'
                j+=1
            nums.append(0)
        nums.sort()
        print(nums)
        # return nums
    similar_count()
    
        



def predict_plitti(demand='趋势预测',YearMonth='202201',Infilename=None,Outdirname=None):
    import datetime
    import itertools
    import os
    # import matplotlib.pyplot as plt
    import numpy as np
    import pandas as pd
    import scipy as sp
    import seaborn as sns  # 热力图
    import statsmodels.api as sm
    import statsmodels.tsa.stattools as ts
    from statsmodels.tsa.arima.model import ARIMA
    def predict():
        # print(Get_data(df,21,1))
        def Get_data(df,year,month,day):
            num=[]
            for i in range(len(df['年'])):
                # print(str(df['年'].iloc[i]),str(df['月'].iloc[i]))
                if df['年'].iloc[i]==year and df['月'].iloc[i]==month:
                    # if df['二级问题'].iloc[i]=='供水' or df['三级问题'].iloc[i]=='供水':
                    if df['日'].iloc[i]==day :
                        num.append(i)
            if num:
                scrdf = pd.DataFrame(df.iloc[num].values)
                scrdf.columns=df.columns
                # print(scrdf)
                return scrdf
            else:
                return np.array([])
        def YearandMonth():
            data20=[]
            data21=[]
            data22=[]
            bigmonth=[1,3,5,7,8,10,12]
            smallmonth=[4,6,9,11]
            specialmonth=[2]
            Year=int(YearMonth[:-2])
            Month=int(YearMonth[-2:])
            data=pd.DataFrame()
            if Year==2020:
                for i in range(Month):
                    if (i+1)!=Month:
                        if (i+1) in bigmonth:
                            for j in range(31):
                                data20.append(Get_data(df,20,i+1,j+1).shape[0])
                        elif (i+1) in smallmonth:
                            for j in range(30):
                                data20.append(Get_data(df,20,i+1,j+1).shape[0])
                        elif (i+1) in specialmonth:
                            for j in range(29):
                                data20.append(Get_data(df,20,i+1,j+1).shape[0])
                    else:
                        for j in range(18):
                            data20.append(Get_data(df,20,i+1,j+1).shape[0])
                data['数值']=data20
            elif Year==2021:
                for i in range(12):
                    if (i+1) in bigmonth:
                        for j in range(31):
                            data20.append(Get_data(df,20,i+1,j+1).shape[0])
                    elif (i+1) in smallmonth:
                        for j in range(30):
                            data20.append(Get_data(df,20,i+1,j+1).shape[0])
                    elif (i+1) in specialmonth:
                        for j in range(29):
                            data20.append(Get_data(df,20,i+1,j+1).shape[0])
                for i in range(Month):
                    if (i+1)!=Month:
                        if (i+1) in bigmonth:
                            for j in range(31):
                                data21.append(Get_data(df,21,i+1,j+1).shape[0])
                        elif (i+1) in smallmonth:
                            for j in range(30):
                                data21.append(Get_data(df,21,i+1,j+1).shape[0])
                        elif (i+1) in specialmonth:
                            for j in range(28):
                                data21.append(Get_data(df,21,i+1,j+1).shape[0])
                    else:
                        for j in range(18):
                            data21.append(Get_data(df,21,i+1,j+1).shape[0])
                data['数值']=data20+data21
            elif Year==2022:
                for i in range(12):
                    if (i+1) in bigmonth:
                        for j in range(31):
                            data20.append(Get_data(df,20,i+1,j+1).shape[0])
                            data21.append(Get_data(df,21,i+1,j+1).shape[0])
                    elif (i+1) in smallmonth:
                        for j in range(30):
                            data20.append(Get_data(df,20,i+1,j+1).shape[0])
                            data21.append(Get_data(df,21,i+1,j+1).shape[0])
                    elif (i+1) in specialmonth:
                        for j in range(29):
                            data20.append(Get_data(df,20,i+1,j+1).shape[0])
                            data21.append(Get_data(df,21,i+1,j).shape[0])
                for i in range(Month):
                    if (i+1)!=Month:
                        if (i+1) in bigmonth:
                            for j in range(31):
                                data22.append(Get_data(df,22,i+1,j+1).shape[0])
                        elif (i+1) in smallmonth:
                            for j in range(30):
                                data22.append(Get_data(df,22,i+1,j+1).shape[0])
                        elif (i+1) in specialmonth:
                            for j in range(28):
                                data22.append(Get_data(df,22,i+1,j+1).shape[0])
                    else:
                        for j in range(18):
                            data22.append(Get_data(df,22,i+1,j+1).shape[0])
                data['数值']=data20+data21+data22

            data.index=pd.Index(pd.date_range(start='20200101',end=YearMonth+'18'))
            
            plt.plot(data.index,data['数值'],'-')
            plt.savefig(Outdirname+'\\pic\\原始数据.jpg')
            # data['数值'].diff(1).plot()
            # plt.savefig(Outdirname+'\\pic\\一阶差分后数据.jpg')
            return data
        
        def judge_stationarity(data):
            dftest = ts.adfuller(data)
            # print(dftest)
            dfoutput = pd.Series(dftest[0:4], index=['Test Statistic','p-value','#Lags Used','Number of Observations Used'])
            stationarity = 1
            for key, value in dftest[4].items():
                dfoutput['Critical Value (%s)'%key] = value 
                if dftest[0] > value:
                        stationarity = 0
            # print(dfoutput)
            print("是否平稳(1/0): %d" %(stationarity))
            return stationarity
        def diff(timeseries,num):
            data_diff = timeseries.diff(num)
            data_diff = data_diff.dropna()
            plt.figure()
            plt.plot(data_diff)
            plt.savefig(Outdirname+'\\pic\\差分后数据.jpg')
            if num==1:
                plt.title('一阶差分')
            elif num==2:
                plt.title('二阶差分')
            # plt.show()
            return data_diff
        def get_P_Q(data):    #获得arim的p值和q值
            pmax = int(5)    #一般阶数不超过 length /10
            qmax = int(5)
            bic_matrix = []
            for p in range(pmax +1):
                temp= []
                for q in range(qmax+1):
                    try:
                        temp.append(ARIMA(data['数值'].diff(1), order=(p, 1, q)).fit().bic)
                    except:
                        temp.append(None)
                    bic_matrix.append(temp)
            # print(bic_matrix)
            bic_matrix = pd.DataFrame(bic_matrix)   #将其转换成Dataframe 数据结构
            p,q = bic_matrix.stack().idxmin()   #先使用stack 展平， 然后使用 idxmin 找出最小值的位置
            # print(u'BIC 最小的p值 和 q 值：%s,%s' %(p,q))  #  BIC 最小的p值 和 q 值：0,1
            return p,q
        df = pd.read_excel(Infilename)
        plt.rcParams['font.sans-serif'] = ['SimHei']
        if not os.path.exists(Outdirname+'\\pic\\'):
            os.makedirs(Outdirname+'\\pic\\')
        data=YearandMonth()
        if judge_stationarity(data)==0:
            myts_diff = diff(data,1)
        if judge_stationarity(myts_diff)==0:
            myts_diff = diff(data,2)
        elif judge_stationarity(myts_diff)==1:
            p,q=get_P_Q(data)
            result_arima = ARIMA(data, order=(p,1,q)).fit()
        # result_arima_fit=model.fit()
            predict_ts = result_arima.predict(start=str(YearMonth)+'19',end=str(int(YearMonth)+1)+'18',typ='levels')  #若不设置typ，预测的值为差分值，需要自己还原
            print(predict_ts)
            sum_predict=np.sum(list(map(int,predict_ts.values.tolist())))
            print(sum_predict)
        # data = data[predict_ts.index]  # 过滤没有预测的记录

        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.figure()
        plt.plot(data.index.tolist(),data['数值'].tolist(),color='skyblue', label='Predict')
        plt.plot(predict_ts.index.tolist(),predict_ts.values.tolist(),color='tomato', label='Original')
        plt.savefig(Outdirname+'\\pic\\预测数据.jpg')

    def plitti():
        df=pd.read_excel(Infilename)
    # if __name__=='__main__':
    if demand=='趋势预测':
        predict()
    elif demand=='突变点标识':
        plitti()

def month_report(Yearmonth=None,Infilename1=None,Infilename2=None,Infilename3=None,Infilename4=None,Outdirname=None):
    Count('直派水务局',Yearmonth,'诉求变化趋势',Infilename2,'./data')
    import codecs
    import os
    import re
    import jieba
    jieba.set_dictionary("./data/dict.txt")
    jieba.initialize()
    # import matplotlib as plot
    import numpy as np
    import pandas as pd
    from django.shortcuts import render
    from docx.shared import Inches, Mm, Pt
    from docxtpl import DocxTemplate, InlineImage
    from numpy.core.numeric import outer
    from pyecharts import options as opts
    from pyecharts.charts import Geo, Map
    from pyecharts.render import make_snapshot
    from snapshot_selenium import snapshot as driver
    Year=int(Yearmonth[:-2])
    Inmonth=int(Yearmonth[-2:])
    class Search_river():
        def __init__(self):
            self.rDict_filename='./data/river.txt' # 河流名称字典
            self.stopkey_filename= './data/stopWord_river.txt'# 停用词字典
            self.tDict_filename='./data/total_dict.txt' # 自建水务词库
        def txt_read(self,files):
            txt_dict = {}
            fopen = open(files,encoding='utf-8')
            for line in fopen.readlines():
                line = str(line).replace("\n","")
                txt_dict[line.split(' ',1)[0]] = line.split(' ',1)[1]
                #split（）函数用法，逗号前面是以什么来分割，后面是分割成n+1个部分，且以数组形式从0开始
            fopen.close()
            return txt_dict


        def get_content(self,df):
            '''
            读取excel中指定两列的内容
            '''
            content1=[]
            content=[]
            for i in range(len(df)):
                content1.append(str(df['诉求内容'].iloc[i]))
                content1.append(str(df['主办单位'].iloc[i]))  
            for j in range(len(content1)):
                if j%2==0:
                    content.append(''.join(content1[j:j+2]))
                else:
                    continue
            return content
        
        def data_clean(self,df):
            '''
            去停用词
            '''
            jieba.load_userdict(self.tDict_filename) # 添加自建词库
            content=self.get_content(df)
            stopkey=[w.strip() for w in codecs.open(self.stopkey_filename, 'r',encoding="utf-8").readlines()]
            m={}
            # 按行边历，先分词，再去除无用词
            for i in range(0,len(content)):
                l=[]
                seg = jieba.lcut(content[i])  
                for j in seg:
                    if j not in stopkey:  
                        l.append(j)
                        m[i]=l
            return m

        def search_river(self,df):
            '''
            在字典中查找河流名称
            '''
            l = []
            suggestion=self.data_clean(df)
            print(suggestion)
            riverdict=self.txt_read(self.rDict_filename)
            keys=list(riverdict.keys())
            # 按行遍历，对于每行中的词，与河流词库匹配。
            for j in range(len(suggestion)):
                    for i in suggestion[j]:
                        count=0
                        for k in riverdict.keys():
                        #  if i in riverdict[k]:
                            c=re.search(i,riverdict[k])
                            a=list(riverdict[k])
                            if c==None or len(i)<2:
                                continue
                            elif c.re.pattern==k :
                                l.append(k)
                                count+=1
                                break
                            elif  a[min(c.span())-1]!=" "  :
                                continue
                            elif a[max(c.span())]==' ':
                                l.append(k)
                                count+=1
                                break  
                        if(count==1):
                            break              
                    if(count==1):
                        continue
                    else:
                        l.append('不详')
            return l

        
    class Pythonword():
        def __init__(self):
            if not os.path.exists(Outdirname+'\\pic\\'):
                os.makedirs(Outdirname+'\\pic\\')
        def Initial(self, df):
            numdf = df['解决类型']
            indexs=[]
            for index, value in enumerate(numdf):
                if '非权属' != value:
                    indexs.append(index)
            scrdf = pd.DataFrame(df.iloc[indexs].values)
            scrdf.columns=df.columns
            return scrdf
        def scrdf(self,df,indexname,judge):#
            df0=df[indexname]
            indexs = []
            for index, value in enumerate(df0):
                if judge == value:
                    indexs.append(index)
            scrdf = pd.DataFrame(df.iloc[indexs].values)
            scrdf.columns=df.columns
            return scrdf

        def getChangeIndex(self,num1,num2,judge):#num1是上个月数据，num2是这个月的数据
            num=[]
            for i in range(len(num2)):
                for j in range(len(num1)):
                    if num2.index[i]==num1.index[j]:
                        if judge==0:
                            if num2.iloc[i]>num1.iloc[j] and (num2.iloc[i]-num1.iloc[j])>0.2:#如果这个月的数据大于上个月，就把这个诉求类型提取,judge=0求取上升的诉求
                                num.append(num2.index[i])
                        elif judge==1:
                            if num2.iloc[i]<num1.iloc[j] and (num1.iloc[j]-num2.iloc[i])>0.2:#judge=1求取下降的诉求
                                num.append(num2.index[i])
            return num
        def getChangeData(self,num1,num2):#num1是上个月数据，num2是这个月的数据
            box=[]
            for i in range(len(num2)):
                for j in range(len(num1)):
                    if num2.index[i]==num1.index[j]:
                        box.append([num2.index[i],num2.iloc[i],num1.iloc[j]])
            df=pd.DataFrame(box)[:10]
            df.columns=['诉求类型', str(int(Inmonth)) + '月数据', str(int(Inmonth) - 1) + '月数据']
            df.set_index('诉求类型',inplace=True)
            return df
        def list2str(self,list0):
            if list0:
                listend=[]
                for i in range(len(list0)):
                    listend.append(list0[i])
                    listend.append('、')
                del(listend[-1])
                str0="".join(listend)
                return str0
            else:
                return None
        def countVar(self,Tdf,Tdf1,Ldf,Ldf1):  # n1是前一个月的数据，n2是后一个月
            num1=[]
            num1.append(Ldf.shape[0])
            num1.append(Ldf.shape[0]-Ldf1.shape[0])
            num1.append(Ldf1.shape[0])
            num2=[]
            num2.append(Tdf.shape[0])
            num2.append(Tdf.shape[0] - Tdf1.shape[0])
            num2.append(Tdf1.shape[0])
            num = []
            for i in range(len(num2)):
                num.append(int(num2[i]))
            if num1[2] > num2[2]:
                num.append(int(num1[2] - num2[2]))
                num.append('下降')
                num.append(round(((num1[2] / num2[2] - 1) * 100), 2))
            else:
                num.append(int(num2[2] - num1[2]))
                num.append('上升')
                num.append(round(((num2[2] / num1[2] - 1) * 100), 2))
            return num

        def getformat2f(self,dfstu):
            count = [float('{:.4f}'.format(i)) for i in dfstu.values]
            count = [round(i * 100, 2) for i in count]
            return count

    class excel2word1(Pythonword):
        def __init__(self,lastmonth,thismonth):
            Pythonword.__init__(self)
            self.key1='诉求类型一级'
            self.key2='诉求类型二级'
            self.key3='诉求类型三级'
            self.keywater='河流名称'
            self.keyname0 = '解决类型'
            self.keyname1 = '非权属'
            self.Tdf = pd.read_excel(thismonth)
            self.Ldf = pd.read_excel(lastmonth)
            river=Search_river().search_river(self.Tdf)
            self.Tdf['河流名称']=river
            self.Tdf=pd.concat([self.Tdf,self.gettime(self.Tdf)],axis=1)
            river1 = Search_river().search_river(self.Ldf)
            self.Ldf['河流名称'] = river1
            self.Ldf = pd.concat([self.Ldf, self.gettime(self.Ldf)],axis=1)
            self.FTdf = self.Initial(self.Tdf)
            self.FLdf = self.Initial(self.Ldf)
            
        def getdict0(self):
            c0 = [f'c{n}' for n in range(6)]
            list=self.countVar(self.Tdf,self.FTdf,self.Ldf,self.FLdf)
            return dict(zip(c0,list)) 
        def getpic1(self):
            pt = self.getChangeData(self.FLdf[self.key1].value_counts(), self.FTdf[self.key1].value_counts())
            plt.rcParams['font.sans-serif'] = ['SimHei']
            pt.plot.bar()
            # plt.bar(pt.index,pt[str(int(Inmonth)) + '月数据'])
            # plt.bar(pt.index,pt[str(int(Inmonth) - 1) + '月数据'])
            x = range(0, len(pt.index.tolist()), 1)
            plt.xticks(x,rotation=60)
            plt.subplots_adjust(bottom=0.35)
            plt.title('直派水务局诉求')
            plt.savefig(Outdirname + '\\pic\\pic0.jpg')
            plt.close()
        def getdict1(self,lastlevel1,thislevel1):
            a0 = [f'a{n}' for n in range(10)]       
            list1 = thislevel1.index.tolist()[:6]
            listc0 = self.getChangeIndex(lastlevel1, thislevel1,0)
            listc1 = self.getChangeIndex(lastlevel1, thislevel1,1)
            list1.append(self.list2str(listc0))
            list1.append(self.list2str(listc1))
            a0dict = dict(zip(a0, list1))
            return a0dict
        def getdict2(self,valuecounts,valuecountsp):
            list0=valuecounts.values.tolist()[:6]
            list1=self.getformat2f(valuecountsp)
            list0.extend(list1)
            a0 = [f'b{n}' for n in range(len(list0))]
            return dict(zip(a0,list0))
        
        def gettime(self,df):
            time=[]
            timestamp=df['登记时间']
            for i in range(len(timestamp)):
                time.append([timestamp.iloc[i][:4],timestamp.iloc[i][5:7],timestamp.iloc[i][8:10],timestamp.iloc[i][11:13]])
            dftime = pd.DataFrame(time)
            dftime.columns = ['年', '月', '日','时']
            return dftime
        def getdicttime(self,cbi):
            listtime=[]
            # listtimeh=[]
            for i in range(len(cbi)):
                listtime.extend(self.get2levelData(cbi[i],'日',3)[:6])
                # listtimeh.extend(self.get2levelData(cbi[i], '时', '时', 2)[:4])
            a0 = [f'time{n}' for n in range(len(listtime))]
            # ah = [f'timeh{n}' for n in range(len(listtimeh))]
            new_dict={}
            new_dict.update(dict(zip(a0, listtime)))
            # new_dict.update(dict(zip(ah, listtimeh)))
            return new_dict
        def get2levelData(self,tmindex,key,num):
            la0=[]
            dfvc0=self.scrdf(self.Tdf,self.key1, tmindex)[key].value_counts()
            if '其它' in dfvc0.index.tolist():
                dfvc0=dfvc0.drop('其它',axis=0)
                # dfvc00=dfvc00.drop('其它',axis=0)
            c0a =dfvc0.index.tolist()[:num]
            c0a.extend('' for _ in range(num-len(c0a)))
            # if key2==key3:
            c0a1 = dfvc0.values.tolist()[:num]
            c0a1.extend('' for _ in range(num-len(c0a1)))
            la0.extend(c0a)
            la0.extend(c0a1)
            # la0.extend(c0a2)
            return la0
        def getleveldata(self,ldf,tdf,judgename):#ldf是上个月的数据，tdf是这个月得数据，judgename是一级诉求里判断条件，二级诉求
            dictlist=[]
            dfv0=self.scrdf(tdf,'诉求类型一级',judgename)['诉求类型二级'].value_counts()
            dfv00=self.scrdf(tdf,'诉求类型一级',judgename)['诉求类型二级'].value_counts(normalize=True)
            if '其它' in dfv0.index.tolist():
                dfv0=dfv0.drop('其它',axis=0)
                dfv00=dfv00.drop('其它',axis=0)
            dfv00=self.getformat2f(dfv00)
            if len(dfv0.index.tolist())>=2:
                dictlist.extend(dfv0.index.tolist()[:2])
                dictlist.extend(dfv0.values.tolist()[:2])
                dictlist.extend(dfv00[:2])
            else:
                dictlist.extend(dfv0.index.tolist())
                dictlist.extend('' for _ in range(2-len(dfv0.index.tolist())))
                dictlist.extend(dfv0.values.tolist())
                dictlist.extend(0 for _ in range(2-len(dfv0.values.tolist())))
                dictlist.extend(dfv00)
                dictlist.extend(0 for _ in range(2-len(dfv00)))
            # print(dictlist)
            dfvl0=self.scrdf(ldf,'诉求类型一级',judgename)['诉求类型二级'].value_counts()

            if len(dfv0.index.tolist())>=1:
                for i in range(len(dfvl0)):
                    if dfv0.index[0]==dfvl0.index[i]:
                        if dfv0.iloc[0]>dfvl0.iloc[i]:
                            dictlist.append('增长')
                            dictlist.append(dfv0.iloc[0]-dfvl0.iloc[i])
                        else:
                            dictlist.append('下降')
                            dictlist.append(dfvl0.iloc[i]-dfv0.iloc[0])
            else:
                dictlist.extend(['',''])
            if len(dfv0.index.tolist())>=2:
                for i in range(len(dfvl0)):
                    if dfv0.index[1]==dfvl0.index[i]:
                        if dfv0.iloc[1]>dfvl0.iloc[i]:
                            dictlist.append('增长')
                            dictlist.append(dfv0.iloc[1]-dfvl0.iloc[i])
                        else:
                            dictlist.append('下降')
                            dictlist.append(dfvl0.iloc[i]-dfv0.iloc[1])
            else:
                dictlist.extend(['',''])
            
            dictlist.append(self.list2str(self.scrdf(tdf,'诉求类型二级',dictlist[0])['诉求类型三级'].value_counts().index.tolist()))
            dictlist.append(self.list2str(self.scrdf(tdf,'诉求类型二级',dictlist[1])['诉求类型三级'].value_counts().index.tolist()))
            # print(dictlist)
            return dictlist
        def cbdw(self,df,cbi,name):
            strlist=[]
            cbdfindex=[]
            if name=='主办单位':          
                cbdf=self.scrdf(df,'诉求类型一级',cbi)[name].value_counts()
                if len(cbdf)>=5:
                    for i in range(5):
                        cbdfindex.append(cbdf.index.tolist()[i])
                        cbdfindex.append('（')
                        cbdfindex.append(str(cbdf.values.tolist()[i]))
                        cbdfindex.append('）件')
                        cbdfindex.append('、')
                    del(cbdfindex[-1])
                    str0="".join(cbdfindex)
                    strlist.append(str0)
                else:
                    for i in range(len(cbdf)):
                        cbdfindex.append(cbdf.index.tolist()[i])
                        cbdfindex.append('（')
                        cbdfindex.append(str(cbdf.values.tolist()[i]))
                        cbdfindex.append('）件')
                        cbdfindex.append('、')
                    del(cbdfindex[-1])
                    str0="".join(cbdfindex)
                    strlist.append(str0)
            elif name=='河流名称':
                cb=self.scrdf(df,'诉求类型一级',cbi)['诉求类型二级'].value_counts().index.tolist()
                if len(cb)<=1:
                    cbdf=self.scrdf(df,'诉求类型二级',cb[0])['河流名称'].value_counts()
                    if len(cbdf)>=5:
                        for i in range(5):
                            cbdfindex.append(cbdf.index.tolist()[i])
                            cbdfindex.append('（')
                            cbdfindex.append(str(cbdf.values.tolist()[i]))
                            cbdfindex.append('）件')
                            cbdfindex.append('、')
                        del(cbdfindex[-1])
                        str0="".join(cbdfindex)
                        strlist.append(str0)
                    else:
                        for i in range(len(cbdf)):
                            cbdfindex.append(cbdf.index.tolist()[i])
                            cbdfindex.append('（')
                            cbdfindex.append(str(cbdf.values.tolist()[i]))
                            cbdfindex.append('）件')
                            cbdfindex.append('、')
                        del(cbdfindex[-1])
                        str0="".join(cbdfindex)
                        strlist.append(str0)
                if len(cb)>=2:
                    cbdf=self.scrdf(df,'诉求类型二级',cb[1])['河流名称'].value_counts()
                    if len(cbdf)>=5:
                        for i in range(5):
                            cbdfindex.append(cbdf.index.tolist()[i])
                            cbdfindex.append('（')
                            cbdfindex.append(str(cbdf.values.tolist()[i]))
                            cbdfindex.append('）件')
                            cbdfindex.append('、')
                        del(cbdfindex[-1])
                        str0="".join(cbdfindex)
                        strlist.append(str0)
                    else:
                        for i in range(len(cbdf)):
                            cbdfindex.append(cbdf.index.tolist()[i])
                            cbdfindex.append('（')
                            cbdfindex.append(str(cbdf.values.tolist()[i]))
                            cbdfindex.append('）件')
                            cbdfindex.append('、')
                        del(cbdfindex[-1])
                        str0="".join(cbdfindex)
                        strlist.append(str0)

            return strlist
        def get_row_data(self,name):
            df=pd.read_excel('./data/诉求风险趋势变化表.xlsx')[-32:]
            ptdf = df.iloc[df['全部诉求'].values.tolist().index(name)]
            datadf = ptdf.values.tolist()
            datadf.remove(name)
            datadf = list(map(int, datadf))
            return datadf
        def compareDict(self,tli):
            a0a0=[]
            for i in range(5):
                monthdata=self.get_row_data(tli[i])
                lastY=monthdata[-13]
                lastM=monthdata[-2]
                thisM=monthdata[-1]
                a0a0.append(self.compareLastMandY(tli[i],lastM,lastY,thisM))
            a0 = [f'aa{n}' for n in range(len(a0a0))]
            return dict(zip(a0,a0a0))
        def compareLastMandY(self,leiname,lastM,lastY,thisM):
            if thisM> lastM and thisM >lastY:
                if lastY!=0:
                    stra0=leiname+'类诉求较上月增加'+str(thisM-lastM)+'件，较去年同期增加'+str(thisM-lastY)+'件，增长率为'+"{:.2f}".format((thisM-lastY)/lastY*100)+'%。'
                else:
                    stra0 = leiname + '类诉求较上月增加' + str(thisM - lastM) + '件，较去年同期增加' + str(thisM - lastY) + '件，下降率为' + 'None。'
            elif thisM> lastM and thisM < lastY:
                if lastY!=0:
                    stra0=leiname+'类诉求较上月增加'+str(thisM-lastM)+'件，较去年同期减少'+str(lastY-thisM)+'件，下降率为'+"{:.2f}".format((lastY-thisM)/lastY*100)+'%。'
                else:
                    stra0 = leiname + '类诉求较上月增加' + str(thisM - lastM) + '件，较去年同期减少' + str(lastY - thisM) + '件，下降率为' +'None。'
            elif thisM < lastM and thisM > lastY:
                if lastY!=0:
                    stra0=leiname+'类诉求较上月减少'+str(lastM-thisM)+'件，较去年同期增加'+str(thisM-lastY)+'件，增长率为'+"{:.2f}".format((thisM-lastY)/lastY*100)+'%。'
                else:
                    stra0 = leiname + '类诉求较上月减少' + str(lastM-thisM) + '件，较去年同期增加' + str(thisM - lastY) + '件，增长率为' +'None。'
            elif thisM < lastM and thisM < lastY:
                if lastY!=0:
                    stra0=leiname+'类诉求较上月减少'+str(lastM-thisM)+'件，较去年同期减少'+str(lastY-thisM)+'件，下降率为'+"{:.2f}".format((lastY-thisM)/lastY*100)+'%。'
                else:
                    stra0 = leiname + '类诉求较上月减少' + str(lastM-thisM) + '件，较去年同期减少' + str(lastY - thisM) + '件，下降率为' +'None。'
            else:
                stra0=None
            return stra0
        def getConstrast(self,name,idx):
            def draw_line_base(month,datadf,yearlist,length):
            # plt.figure()
                plt.plot()
                step = int(length/12)
                # print(length,step)
                for i in range(step):
                    # print("%d",i*12)
                    plt.plot(month[:12],datadf[int(i*12):int((i+1)*12)], marker='o', markersize=3)
                    for a, b in zip(month[:12], datadf[int(i*12):int((i+1)*12)]):
                        plt.text(a, b, b, ha='center', va='bottom', fontsize=10)  # 设置数据标签位置及大小
                plt.plot(month[:length-step*12],datadf[step*12:length],marker='o',markersize=3)
                for a, b in zip(month[:length-step*12], datadf[step*12:length]):
                    plt.text(a, b, b, ha='center', va='bottom', fontsize=10)
                plt.legend(yearlist[:step+1])  # 设置折线名称
            # plt.show()
            df=pd.read_excel('./data/诉求风险趋势变化表.xlsx')[-32:]
            plt.rcParams['font.sans-serif'] = ['SimHei']
            ptdf = df.iloc[df['全部诉求'].values.tolist().index(name)]
            datadf = ptdf.values.tolist()
            datadf.remove(name)
            datadf = list(map(int, datadf))
            yearlist=['2020年','2021年','2022年','2023年','2024年','2025年','2026年','2027年','2028年']
            month = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']
            length=len(datadf)
            draw_line_base(month,datadf,yearlist,length)
            plt.title(label=name)
            plt.xlabel('时间')  # x轴标题
            plt.ylabel('诉求值')  # y轴标题
            plt.savefig(Outdirname  + '\\pic\\pica'+ str(idx) + '.jpg')
            plt.close()

        def getWordmain(self):
            self.getpic1()
            thislevel1 = self.FTdf[self.key1].value_counts().drop('其它',axis=0)
            thislevel1p = self.FTdf[self.key1].value_counts(normalize=True).drop('其它',axis=0)
            lastlevel1 = self.FLdf[self.key1].value_counts().drop('其它',axis=0)
            thislevel1i = thislevel1.index.tolist()
            list0=[]
            list1=[]
            list2=[]
            for i in range(5):
                list0.extend(self.getleveldata(self.FLdf,self.FTdf,thislevel1i[i]))
                self.getConstrast(thislevel1i[i],i)
                list1.extend(self.cbdw(self.FTdf,thislevel1i[i],'河流名称'))
                list2.extend(self.cbdw(self.FTdf,thislevel1i[i],'主办单位'))
                
            # dict0=zip(dict([f'a0{n}' for n in range(len(list0))],list0))
            new_dict = {}
            # new_dict.update(self.getMonthTime())
            new_dict.update(self.getdict0())
            new_dict.update(dict(zip([f'a0{n}' for n in range(len(list0))],list0)))
            new_dict.update(dict(zip([f'ab{n}' for n in range(len(list1))],list1)))
            new_dict.update(dict(zip([f'ac{n}' for n in range(len(list2))],list2)))
            new_dict.update(self.getdict1(lastlevel1, thislevel1))
            new_dict.update(self.getdict2(thislevel1, thislevel1p))
            new_dict.update(self.compareDict(thislevel1i))
            new_dict.update(self.getdicttime(thislevel1i))
            return new_dict

    class excel2word2(Pythonword):
        def __init__(self,Lmonth,Tmonth):
            Pythonword.__init__(self)
            self.tdf = pd.read_excel(Tmonth,sheet_name='区街道乡镇')
            self.ldf = pd.read_excel(Lmonth,sheet_name='区街道乡镇')
            self.cbcounts = self.tdf['承办单位'].value_counts()


        def getdict1(self):
            chengban1 = '北京市自来水集团有限责任公司'
            chengban2 = '北京城市排水集团有限责任公司'
            d0=[]
            d0.append(self.tdf.shape[0])
            if chengban1 in self.cbcounts.index.tolist():
                if chengban2 in self.cbcounts.index.tolist():
                    d0.append(self.cbcounts.loc[chengban1])
                    d0.append(self.cbcounts.loc[chengban2])
            listd0 = [f'd{n}' for n in range(3)]
            dict1 = dict(zip(listd0,d0))
            return dict1
        def getdict2(self,cbs,cbp):
            cbi = cbs.index.tolist()[:10]
            cbv=cbs.values.tolist()[:10]
            cbi.extend(cbv)
            list1 = self.getformat2f(cbp)
            cbi.extend(list1)
            listd00=[f'd0{n}' for n in range(len(cbi))]
            dict2=dict(zip(listd00,cbi))
            return dict2
        def getpic2(self,pt):
            plt.rcParams['font.sans-serif'] = ['SimHei']
            pt.plot.bar()
            # plt.bar(pt.index,pt[str(int(Inmonth)) + '月数据'])
            # plt.bar(pt.index,pt[str(int(Inmonth) - 1) + '月数据'])
            x = range(0, len(pt.index.tolist()), 1)
            plt.xticks(x,rotation=60)
            plt.subplots_adjust(bottom=0.35)
            plt.title('直派区街道乡镇诉求')
            plt.savefig(Outdirname +'\\pic\\pic1.jpg')
        def getdict3(self):
            num=[]
            num.append(self.list2str(self.getChangeIndex(self.ldf['三级问题'].value_counts(),self.tdf['三级问题'].value_counts(),0)))
            num.append(self.list2str(self.getChangeIndex(self.ldf['三级问题'].value_counts(),self.tdf['三级问题'].value_counts(),1)))
            listd0 = [f'd00{n}' for n in range(2)]
            dict3=dict(zip(listd0,num))
            return dict3
        def getpic(self,num0,num,cbi):
            pt = self.getChangeData(self.scrdf(self.ldf, '三级问题', cbi)['被反映区'].value_counts(),self.scrdf(self.tdf, '三级问题', cbi)['被反映区'].value_counts())
            plt.rcParams['font.sans-serif'] = ['SimHei']
            pt.plot.bar()
            # plt.bar(pt.index,pt[str(int(Inmonth)) + '月数据'])
            # plt.bar(pt.index,pt[str(int(Inmonth) - 1) + '月数据'])
            x = range(0, len(pt.index.tolist()), 1)
            plt.xticks(x, rotation=45)
            plt.subplots_adjust(bottom=0.35)
            plt.title(cbi)
            plt.savefig(Outdirname+ '\\pic\\pic'+str(num0+2)+'.jpg')
            quname=['密云区','延庆区','朝阳区','丰台区','石景山区','海淀区','门头沟区','房山区','通州区','顺义区','昌平区','大兴区','怀柔区','平谷区','东城区','西城区']
            ptindex=self.scrdf(self.tdf, '三级问题', cbi)['被反映区'].value_counts().index.tolist()
            # ptindex=list(set(ptindex+quname))
            ptindex=list((ptindex+ [item for item in quname if str(item) not in ptindex]))
            ptvalue=self.scrdf(self.tdf, '三级问题', cbi)['被反映区'].value_counts().values.tolist()
            if len(ptindex)!=len(ptvalue):
                ptvalue.extend(0 for _ in range(len(ptindex)-len(ptvalue)))
            maxvalue=max(map(int,ptvalue))
            while (maxvalue %5):
                maxvalue+=1
            c = (
                Map()
                .add("", [list(z) for z in zip(ptindex, ptvalue)], "北京",zoom=1.25,aspect_scale=0.9,layout_center=['60%','60%'],
                label_opts=opts.LabelOpts(is_show=True,position='outsideleft',font_size=10,color='#708069',rotate = '30',horizontal_align = 'center',font_weight='bold',vertical_align ='middle'))
                .set_global_opts(
                title_opts=opts.TitleOpts(title=cbi+'类诉求'),
                visualmap_opts=opts.VisualMapOpts(type_='color', orient='vertical',max_=maxvalue,
                                                    is_piecewise=True,range_color=['#71ae9b','#8bb08e', '#c1d389','#e3d76b','#f7b44a','#f27127','#cc8775']
                ))
            )
            make_snapshot(driver, c.render(), Outdirname +'\\pic\\picb'+ str(num)+".png")
            b = (
                Map()
                .add("", [list(z) for z in zip(ptindex, ptvalue)], "北京",zoom=1.25,aspect_scale=0.9,layout_center=['60%','60%'],
                label_opts=opts.LabelOpts(is_show=True,position='outsideleft',font_size=10,color='#708069',rotate = '30',horizontal_align = 'center',font_weight='bold',vertical_align ='middle'))
                .set_global_opts(
                title_opts=opts.TitleOpts(title=cbi+'类诉求'),
                visualmap_opts=opts.VisualMapOpts(type_='color', orient='vertical',is_piecewise=True,pieces=[{'min':0,'max':10,'label':'[0-10)',"color":'#71ae9b'},{'min':11,'max':30,'label':'[10-30)',"color":'#8bb08e'},{'min':31,'max':100,'label':'[30-100)',"color":'#c1d389'},{'min':101,'max':200,'label':'[100-200)',"color":'#e3d76b'},{'min':201,'max':350,'label':'[200-350)',"color":'#f7b44a'},
                                                            {'min':350,'label':'[350,)',"color":'#f27127'}]
                                                            # {'min':0,'max'=10,label:'0-10','#cc8775']
                                                    )
            ))
            make_snapshot(driver, b.render(), Outdirname +'\\pic\\picb'+ str(num+1)+".png")
            

        def get1level(self,indexname,var,key1,key2,num):
            lc0a = [str(var)+f'{n}' for n in range(3*num)]
            c0a = self.get2levelData(indexname, key1, key2,num)
            dict1 = dict(zip(lc0a, c0a))
            return dict1

        def get2levelData(self,tmindex,key1,key2,num):
            la0=[]
            c0a = self.scrdf(self.tdf,key1, tmindex)[key2].value_counts().index.tolist()[:num]
            c0a.extend('' for _ in range(num-len(c0a)))
            c0a1 = self.scrdf(self.tdf,key1, tmindex)[key2].value_counts().values.tolist()[:num]
            c0a1.extend('' for _ in range(num-len(c0a1)))
            c0a2 = self.getformat2f(self.scrdf(self.tdf,key1, tmindex)[key2].value_counts(normalize=True))
            la0.extend(c0a)
            la0.extend(c0a1)
            la0.extend(c0a2)
            return la0

        def getdict4(self,num1,num2):#num1是上个月的数据
            num=[]
            for i in num2.index.tolist():
                for j in num1.index.tolist():
                    if i==j:
                        if num1.loc[j] > num2.loc[i]:
                            num.append('下降')
                            num.append(int(num1.loc[i] - num2.loc[i]))
                        else:
                            num.append('上升')
                            num.append(int(num2.loc[i] - num1.loc[i]))
            listdd = [f'd0d{n}' for n in range(len(num))]
            return dict(zip(listdd,num))
        def getdict51(self,cbi):
            listdict=[]
            vc1=self.scrdf(self.ldf, '三级问题', cbi)['被反映区'].value_counts()
            vc2=self.scrdf(self.tdf, '三级问题', cbi)['被反映区'].value_counts()

            num0=self.getChangeIndex(vc1,vc2,0)
            num1=self.getChangeIndex(vc1,vc2,1)
            listdict.append(self.list2str(num0))
            listdict.append(self.list2str(num1))
            return listdict
        def getdict5(self,cbi):
            num=[]
            for i in range(len(cbi)):
                num.extend(self.getdict51(cbi[i]))
            listdd = [f'd0a{n}' for n in range(len(num))]
            return dict(zip(listdd, num))
        def getWordmain(self):
            name = '三级问题'
            key1='被反映区'
            cbcounts = self.tdf[name].value_counts()
            cbi=cbcounts.index.tolist()
            cbcountsp = self.tdf[name].value_counts(normalize=True)
            pt = self.getChangeData(self.ldf[name].value_counts(),
                                    self.tdf[name].value_counts())#三级问题获得counts
            self.getpic2(pt)
            num=[0,2,4,6,8,10,12,14,16,18]
            for i in range(10):
                self.getpic(i,num[i],cbi[i])
            new_dict={}
            new_dict.update(self.getdict1())
            new_dict.update(self.getdict2(cbcounts,cbcountsp))
            new_dict.update(self.getdict3())
            new_dict.update(self.get1level(cbi[0],'da',name,key1,5))
            new_dict.update(self.get1level(cbi[1], 'db', name, key1, 5))
            new_dict.update(self.get1level(cbi[2], 'dc', name, key1, 5))
            new_dict.update(self.get1level(cbi[3], 'dd', name, key1, 5))
            new_dict.update(self.get1level(cbi[4], 'de', name, key1, 5))
            new_dict.update(self.get1level(cbi[5], 'df', name, key1, 5))
            new_dict.update(self.get1level(cbi[6], 'dg', name, key1, 5))
            new_dict.update(self.get1level(cbi[7], 'dh', name, key1, 5))
            new_dict.update(self.get1level(cbi[8], 'di', name, key1, 5))
            new_dict.update(self.get1level(cbi[9], 'dj', name, key1, 5))
            new_dict.update(self.getdict4(self.ldf[name].value_counts(),
                                self.tdf[name].value_counts()))
            new_dict.update(self.getdict5(cbi[:10]))
            return new_dict

    def getMonthTime(month):
            time = ['t0', 't1', 't2', 't3']
            month2 = ['1', '19', '2', '18']
            month3 = ['2', '19', '3', '18']
            month4 = ['3', '19', '4', '18']
            month5 = ['4', '19', '5', '18']
            month6 = ['5', '19', '6', '18']
            month7 = ['6', '19', '7', '18']
            month8 = ['7', '19', '8', '18']
            month9 = ['8', '19', '9', '18']
            month10 = ['9', '19', '10', '18']
            month11 = ['10', '19', '11', '18']
            month12 = ['11', '19', '12', '18']
            month1 = ['12', '19', '1', '18']
            if month == 1:
                month_dict = dict(zip(time, month1))
                return month_dict
            elif month == 2:
                month_dict = dict(zip(time, month2))
                return month_dict
            elif month == 3:
                month_dict = dict(zip(time, month3))
                return month_dict
            elif month == 4:
                month_dict = dict(zip(time, month4))
                return month_dict
            elif month == 5:
                month_dict = dict(zip(time, month5))
                return month_dict
            elif month == 6:
                month_dict = dict(zip(time, month6))
                return month_dict
            elif month == 7:
                month_dict = dict(zip(time, month7))
                return month_dict
            elif month == 8:
                month_dict = dict(zip(time, month8))
                return month_dict
            elif month == 9:
                month_dict = dict(zip(time, month9))
                return month_dict
            elif month == 10:
                month_dict = dict(zip(time, month10))
                return month_dict
            elif month == 11:
                month_dict = dict(zip(time, month11))
                return month_dict
            elif month == 12:
                month_dict = dict(zip(time, month12))
                return month_dict
            else:
                return None
                
                # if __name__=='__main__':
    import time
    asset_url = './data/模板.docx'
    context1=excel2word1(Infilename1,Infilename2).getWordmain()
    context2=excel2word2(Infilename3,Infilename4).getWordmain()
    tpl = DocxTemplate(asset_url)
    picture={}
    for i in range(12):
        picture.update({'pic'+str(i): InlineImage(tpl, Outdirname+'\\pic\\pic'+str(i)+'.jpg', width=Mm(130), height=Mm(100))})
    for i in range(5):
        picture.update({'pica' + str(i): InlineImage(tpl, Outdirname+'\\pic\\pica'+str(i) + '.jpg', width=Mm(130), height=Mm(100))})
    for i in range(20):
        picture.update({'picb' + str(i): InlineImage(tpl, Outdirname+'\\pic\\picb'+str(i) + '.png', width=Mm(160), height=Mm(130))})
    
    # print(context2)

    context1.update(getMonthTime(int(Inmonth)))
    context1.update(picture)
    context1.update(context2)
    tpl.render(context1)
    tpl.save(Outdirname+'\\涉水数据' + str(int(Inmonth))+'月份数据分析报告'+time.strftime("%Y%m%d")+'.docx')

@Gooey(
richtext_controls=True,                 # 打开终端对颜色支持
program_name="市水务局涉水诉求工单数据分析",        # 程序名称
encoding="utf-8",                  # 设置编码格式，打包的时候遇到问题
progress_regex=r"^progress: (\d+)%$",   # 正则，用于模式化运行时进度信息
default_size=(800,600)
)
def start():
    parser = GooeyParser(description='')

    subs = parser.add_subparsers(help='commands', dest='command')

    Fir_parser = subs.add_parser('属性增补')
    
    Fir_parser.add_argument (
            "data_type", 
    metavar='数据源格式',
    help=   "请选择数据源格式",
    choices=['直派水务局','直派行业'], 
    default='直派水务局')
    Fir_parser.add_argument (
            "device_type",
    metavar='增补属性',
    help=   "请选择需要增补的数据",
    choices=['河湖','小区（村）','供水属性标注','汇总数据以及时间分列'],
    default='河湖')

    Fir_parser.add_argument (
            "FirInFilename",
    metavar='选择文件',
    widget='MultiFileChooser',
    default=None,
    help='请选择源数据')
    
    Fir_parser.add_argument(
            "FirOutDirname", 
    metavar='输出文件目录',
    widget='DirChooser',
    default=None,
    help='请选择输出文件目录')
    

    Sec_parser = subs.add_parser('智能分类和原因提取')
    Sec_parser.add_argument(
            "classify_reason", 
    metavar='智能分类和原因提取',
    choices=['诉求智能分类（直派水务局）','诉求原因提取（直派行业）'], 
    default=None)
    Sec_parser.add_argument (
            "SecInFilename",
    metavar='选择输入文件',
    widget='MultiFileChooser',
    default=None,
    help='请选择源数据')
    Sec_parser.add_argument (
            "SecOutDirname",
    metavar='保存文件',
    widget='DirChooser',
    default=None,
    help='请选择保存文件目录')
    
    third_parser=subs.add_parser('统计分析')
    third_parser.add_argument (
            "data_type", 
    metavar='数据源格式',
    help=   "请选择数据源格式",
    choices=['直派水务局','直派行业'], 
    default='直派水务局')
    third_parser.add_argument(
        "YearMonth",
    metavar='输入当前年月',
    widget='TextField',
    default='202101',
    help='')
    third_parser.add_argument (
            "Count",
    metavar='统计',
    help=   "请选择需要统计的数据",
    choices=['单位解决类型','诉求变化趋势','水利工程运行维护','各区街道集团','只绘图'],
    default='单位解决类型')
    third_parser.add_argument (
            "Draw",
    metavar='绘图',
    help=   "请选择需要绘制的数据",
    choices=['水利工程运行维护变化趋势-折线图','水环境维护变化趋势-折线图','雨水与海绵城市变化趋势-折线图','饼状图','柱状图（直派水务局一级诉求）','柱状图（直派行业三级问题）','地理分布图（区）','只统计'],
    default='水利工程运行维护变化趋势-折线图')
    third_parser.add_argument (
            "ThirdInFilename",
    metavar='输入文件',
    widget='MultiFileChooser',
    default=None,
    help='请选择源数据')
    third_parser.add_argument (
            "ThirdOutDirname",
    metavar='保存文件',
    widget='DirChooser',
    default=None,
    help='请选择保存文件目录')
    fouth_parser=subs.add_parser('相似性分析、趋势预测和突变点标识')
    fouth_parser.add_argument (
            "demand", 
    metavar='需求',
    help=   "请选择需求",
    choices=['相似性分析','趋势预测','突变点标识'], 
    default='相似性分析')
    fouth_parser.add_argument(
        "YearMonth",
    metavar='输入当前年月',
    widget='TextField',
    default='202001',
    help='')
    fouth_parser.add_argument(
        "FouthInFilename",
    metavar='输入Excel文件',
    widget='MultiFileChooser',
    default=None,
    help='请选择输入文件（汇总excel）')
    
    fouth_parser.add_argument(
            "FouthOutDirname",
    metavar='保存文件',
    widget='DirChooser',
    default=None,
    help='请选择保存文件目录')

    fifth_parser = subs.add_parser('月报')
    fifth_parser.add_argument(
        "Yearmonth",
    metavar='输入当前年月',
    widget='TextField',
    default='202001',
    help='')
    fifth_parser.add_argument(
        "FifthInFilename1",
    metavar='输入文件',
    widget='MultiFileChooser',
    default=None,
    help='请输入上月直派水务局数据')
    fifth_parser.add_argument(
        "FifthInFilename2",
    metavar='输入文件',
    widget='MultiFileChooser',
    default=None,
    help='请输入该月直派水务局数据')
    fifth_parser.add_argument(
        "FifthInFilename3",
    metavar='输入文件',
    widget='MultiFileChooser',
    default=None,
    help='请输入上月直派行业数据')
    fifth_parser.add_argument(
        "FifthInFilename4",
    metavar='输入文件',
    widget='MultiFileChooser',
    default=None,
    help='请输入该月直派行业数据')
    fifth_parser.add_argument(
            "FifthOutDirname",
    metavar='保存文件',
    widget='DirChooser',
    default=None,
    help='请选择保存文件目录')

    args = parser.parse_args()
    print(args,flush=True)    # 坑点：flush=True在打包的时候会用到
   	# 将界面收集的参数进行处理
    return args
if __name__ == '__main__':

    args=start()#InFilename='输入单文件',InDirname='输入文件目录',OutDirname='输出文件目录'
    if args.command=='属性增补':
        if args.device_type=='河湖':
            Extract_river(args.data_type,args.FirInFilename,args.FirOutDirname)
        elif args.device_type=='小区（村）':
            Extract_village(args.data_type,args.FirInFilename,args.FirOutDirname)
        elif args.device_type=='供水属性标注':
            Extract_attribute(args.FirInFilename,args.FirOutDirname)
        elif args.device_type=='汇总数据以及时间分列':
            Sum_Extract_time(args.data_type,args.FirInFilename,args.FirOutDirname)

    
    elif args.command=='智能分类和原因提取':
        if args.classify_reason=='直派水务局':
            Classification(args.SecInFilename,args.SecOutDirname)

        elif args.classify_reason=='直派行业':
            Extract_reason(args.SecInFilename,args.SecOutDirname)

    elif args.command=='统计分析':
        if args.Draw=='只统计':
            Count(args.data_type,args.YearMonth,args.Count,args.ThirdInFilename,args.ThirdOutDirname)

        elif args.Count=='只绘图':
            Draw(args.data_type,args.Draw,args.ThirdInFilename,args.ThirdOutDirname)

    elif args.command=='相似性分析、趋势预测和突变点标识':
        if args.demand=='相似性分析':
            similar_ayalyse(args.YearMonth,args.FouthInFilename,args.FouthOutDirname)
        else:
            predict_plitti(args.demand,args.YearMonth,args.FouthInFilename,args.FouthOutDirname)

    elif args.command=='月报':
        month_report(args.Yearmonth,args.FifthInFilename1,args.FifthInFilename2,args.FifthInFilename3,args.FifthInFilename4,args.FifthOutDirname)

