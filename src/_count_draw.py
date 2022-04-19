def Count(data_type='直派水务局',YearMonth=None,count_type=None,InDirname=None,Outdirname=None):
    import operator
    import os
    from collections import Counter

    import openpyxl
    import pandas as pd
    from openpyxl import load_workbook

    filepath1 = 'demo_excel/办公单位.xlsx'
    filepath2 = 'demo_excel/诉求类型一级表.xlsx'
    # filepath3 = './tongji_out/统计报告.xlsx'
    # filepath4 = './tongji_out/诉求风险趋势变化表.xlsx'
    filepath3 = 'demo_excel/out_demo/统计报告.xlsx'
    filepath4 = 'demo_excel/out_demo/诉求风险趋势变化表.xlsx'
    filepath5 = 'demo_excel/街道二级诉求.xlsx'
    filepath6 = 'demo_excel/各区街道集团统计样板.xlsx'
    filepath7 = Outdirname + '\\各区街道集团统计.xlsx'
    filepath8 = Outdirname + '\\统计报告.xlsx'
    filepath9 = Outdirname + '\\诉求风险趋势变化表.xlsx'
    year=int(YearMonth[:-2])
    month=int(YearMonth[-2:])
    # filepath7 = Outdirname+'./town_out/各区街道集团统计.xlsx'
    # filepath8 = './demo_excel/汇总样板.xlsx'    #createxcel 的filepath3、4
    # filepath9 = './demo_excel/诉求类型样板.xlsx'

    filelist1 = ['2021.1','2021.2','2021.3','2021.4','2021.5','2021.6','2021.7','2021.8','2021.9','2021.10','2021.11','2021.12','2022.1','2022.2','2022.3','2022.4','2022.5','2022.6','2022.7','2022.8','2022.9','2022.10','2022.11','2022.12','2023.1','2023.2','2023.3','2023.4','2023.5','2023.6','2023.7','2023.8','2023.9','2023.10','2023.11','2023.12']
    if year==2020:
        filelist1=filelist1[:month]
    elif year==2021:
        filelist1=filelist1[:12+month]
    elif year==2022:
        filelist1=filelist1[:24+month]
    elif year==2023:
        filelist1=filelist1[:36+month]
    excel_num1 = len(filelist1)
    # print(filelist, excel_num,filelist[excel_num-3].replace('.xlsx','') )  # 这里filelist从0开始，到（excel数量-1）结束

    def getexcel_town(choice, excel_filepath, dict_filepath):
        wb = openpyxl.Workbook()  # 创建一个新的excel
        we = wb['Sheet']
        if choice == '各区街道集团':
            we.title = '统计'
            we['A1'] = '各区街道集团'
            we['A2'] = '二级问题'
        month_num = excel_num1
        for i in range(0, month_num):
            we.cell(2, 2 + i).value = filelist1[i]
            we.cell(2, month_num + 2).value = '合计'
            i += 1
        for j in range(0, len(getdict(dict_filepath))):
            we.cell(3 + j, 1).value = getdict(dict_filepath)[j]['名称']
            j += 1

        first_num = j + 4
        if choice == '各区街道集团':
            we.cell(first_num, 1).value = '三级问题（排名降序）'  #
        # second_num=first_num+j+3                   #  已解决+剔除+挂账 部分
        wb.save(excel_filepath)  # 保存表格
        return (first_num)

    def alltongji_town(choice, excel_start, first_path, second_path, third_path):  #### 程序运行时不要打开用到的 Excel ！

        rb = openpyxl.load_workbook(first_path)
        if choice == '各区街道集团':
            wl = rb['统计']
        for i in range(excel_start - 1, excel_num1):  # 决定统计表格起始    *********************************************
            if choice == '各区街道集团':
                new_num = getexcel_town('各区街道集团', first_path, second_path)
            # data = pd.read_excel('./data_town/' + filelist1[i], sheet_name='各区街道集团')
            data = pd.read_excel(InDirname, sheet_name='各区街道集团')
            row = len(data)  # 表格行数
            x_pandas_list = data['二级问题']  # 专业情况
            c = Counter(x_pandas_list)
            # print(c)
            list_len = len(c)
            factor_len = len(getdict(second_path))  # 办理单位数量
            # print(factor_len)
            print(f"{choice}{filelist1[i]}开始统计···")
            all_tongji, not_tongji, other_tongji, type_all_tongji, type_other_tongji = 0, 0, 0, 0, 0
            all_list = {}
            not_list = {}
            other_list = {}
            type_all_list = {}
            type_other_list = {}  # 存储每月各单位的各类型数量
            kong_num = 0  # 办公单位为空
            for j in range(0, factor_len):  # 循环所有办公单位
                all_num = 0  # 全部类型  非权属  已解决+剔除+挂账 （等）
                all_data = []
                for k in range(0, list_len):
                    exit_flag = 0
                    if choice == '各区街道集团':
                        if format(c.most_common()[k][0].replace('（', '').replace('）', '')) == format(
                                getdict(second_path)[j]['名称'].replace('(', '').replace(')',
                                                                                       '')):  # 二级问题所在字典  一个去中文（） ,一个去英文()
                            exit_flag = 1
                            if exit_flag == 1:
                                all_num = c.most_common()[k][1]
                            if exit_flag == 0:
                                all_num = 0
                ##  结果存入数组
                all_list[j] = all_num
                all_data.append(all_list)
                all_tongji += all_num
                # print(all_num, all_tongji, row ,j)
            for j in range(0, factor_len):
                if choice == '各区街道集团':
                    #   数量
                    wl.cell(3 + j, 2 + i).value = int_str(all_data[0][j])  # 统计结果写入 汇总表
            #   保存
            rb.save(third_path)

            if all_tongji != row:
                if choice == '各区街道集团':
                    print(
                        f"此月份表格有{row - all_tongji}条二级问题信息未添加到‘街道二级诉求.xlsx’文件中，或有的名字打错！！！")  # 这里格式要求严格，如果名字没错也没有空的，那么注意名字前后的 ‘ ’，即，空格！
            print(f"{filelist1[i]}统计完成\n")

        dic = []
        for j in range(0, factor_len):
            sum_number = 0
            for k in range(excel_start - 1, excel_num1):
                if choice == '各区街道集团':
                    sum_number += int(wl.cell(3 + j, 2 + k).value)  # 统计结果写入 汇总表
            wl.cell(3 + j, 2 + excel_num1).value = sum_number
            dic.append({'name': getdict(second_path)[j]['名称'], 'number': sum_number})
        rb.save(third_path)
        sorted_x = sorted(dic, key=operator.itemgetter('number'), reverse=True)
        print(f"{sorted_x} ,{len(sorted_x)}\n")

        #### 二级问题排序，创建动态表格框架
        first_num = new_num + 1
        for top_num in range(0, len(sorted_x)):
            wl['A' + int_str(new_num + 1)] = f"第{top_num + 1}位:" + sorted_x[top_num]['name']
            if sorted_x[top_num]['name'] == '中水':
                zs_num = new_num + 3
            if sorted_x[top_num]['name'] == '供水':
                gs_num = new_num + 7
            for top_i in range(0, excel_num1):
                wl.cell(new_num + 1, 2 + top_i).value = filelist1[top_i].replace('.xlsx', '')
                top_i += 1
            town_getdict_len = len(town_getdict(sorted_x[top_num]['name']))  # mou 某二级问题问题的三级字典长度
            for top_j in range(0, town_getdict_len):
                wl.cell(new_num + 2 + top_j, 1).value = town_getdict(sorted_x[top_num]['name'])[top_j]['名称']
                top_j += 1
            new_num += top_j + 2
        # 保存
        rb.save(third_path)
        ##################################################################################
        #### 开始前十类写入
        for top_num in range(0, len(sorted_x)):
            for i in range(excel_start - 1, excel_num1):
                # data = pd.read_excel('./data_town/' + filelist1[i], sheet_name='各区街道集团')
                data = pd.read_excel(InDirname, sheet_name='区街道乡镇')
                # row = len(data)  # 表格行数
                x_pandas_list = data['三级问题']  # 专业情况
                c = Counter(x_pandas_list)
                list_len = len(c)
                factor_len = len(town_getdict(sorted_x[top_num]['name']))  # 三级问题数量
                print(f"第{top_num + 1}位{sorted_x[top_num]['name']}{filelist1[i]}开始统计···")
                all_tongji = 0
                all_list = {}
                for j in range(0, factor_len):  # 循环所有办公单位
                    all_num, not_num, other_num = 0, 0, 0  # 全部类型  非权属  已解决+剔除+挂账 （等）
                    all_data = []
                    for k in range(0, list_len):
                        exit_flag = 0
                        if choice == '各区街道集团':
                            if c.most_common()[k][0] == town_getdict(sorted_x[top_num]['name'])[j]['名称']:
                                exit_flag = 1
                                if exit_flag == 1:
                                    all_num = c.most_common()[k][1]
                                if exit_flag == 0:
                                    all_num = 0
                        ##  结果存入数组
                    all_list[j] = all_num
                    all_data.append(all_list)
                    all_tongji += all_num
                    # print(all_num, all_tongji, row ,j)
                for j in range(0, factor_len):
                    if choice == '各区街道集团':
                        #   数量
                        wl.cell(first_num + 1 + j, 2 + i).value = int_str(all_data[0][j])  # 统计结果写入 汇总表
                ###  找到具有共同三级的二级问题的类的所在  行
                for find_i in range(0, len(getdict(second_path))):
                    if wl.cell(row=find_i + 3, column=1).value == sorted_x[top_num]['name']:
                        the_row = find_i + 3
                        # print(the_row,excel_start,wl.cell(row=the_row,column=i+2 ).value,type(wl.cell(row=the_row,column= i+2).value),all_tongji )
                ####  修改冗余部分
                if all_tongji > int(wl.cell(row=the_row, column=i + 2).value):
                    if sorted_x[top_num]['name'] == '中水':
                        error_num = all_tongji - int(
                            wl.cell(row=the_row, column=i + 2).value)  # 因有共同三级问题情况多出来的   目前只知道缴费
                        if choice == '各区街道集团':
                            wl.cell(zs_num, 2 + i).value = int_str(all_data[0][1] - error_num)  # 统计结果写入 汇总表
                            all_tongji = all_tongji - error_num
                            # print(the_row,wl.cell(row=the_row,column=i+2 ).value,all_data [0][1],all_tongji,error_num)
                    if sorted_x[top_num]['name'] == '供水':
                        error_num = all_tongji - int(
                            wl.cell(row=the_row, column=i + 2).value)  # 因有共同三级问题情况多出来的   目前只知道缴费
                        if choice == '各区街道集团':
                            wl.cell(gs_num, 2 + i).value = int_str(all_data[0][5] - error_num)  # 统计结果写入 汇总表
                            all_tongji = all_tongji - error_num
                            # print(the_row,wl.cell(row=the_row,column=i+2 ).value,all_data[0][5],all_tongji ,error_num)
                #   保存
                rb.save(third_path)

                if all_tongji == int(wl.cell(row=the_row, column=i + 2).value):
                    if choice == '各区街道集团':
                        print(f"此月份表格有{all_tongji}条{sorted_x[top_num]['name']}三级问题信息。")
                #  再次判断是否有共同三级问题情况
                if all_tongji > int(wl.cell(row=the_row, column=i + 2).value):
                    if choice == '各区街道集团':
                        print(
                            f"此月份表格有{all_tongji - int(wl.cell(row=the_row, column=i + 2).value)}条三级问题信息在其他二级问题的下属三级问题中也存在！！")
                if all_tongji < int(wl.cell(row=the_row, column=i + 2).value):  # sorted_x[top_num]['number']:
                    if choice == '各区街道集团':
                        print(f"此月份表格有{int(wl.cell(row=the_row, column=i + 2).value) - all_tongji}条三级问题信息未被统计！！！")
                print(f"{sorted_x[top_num]['name']}{filelist1[i]}统计完成\n")
            first_num += 2 + factor_len
        print('统计完成！')

    def getexcel(choice, excel_filepath, dict_filepath):
        # list_month = ['2020.1', '2020.2', '2020.3', '2020.4', '2020.5', '2020.6', '2020.7', '2020.8', '2020.9',
        #               '2020.10', '2020.11', '2020.12', '2021.1', '2021.2', '2021.3', '2021.4', '2021.5', '2021.6',
        #               '2021.7', '2021.8', '2021.9', '2021.10', '2021.11', '2021.12']

        wb = openpyxl.Workbook()  # 创建一个新的excel
        we = wb['Sheet']
        we2 = wb.create_sheet('sheet2', 1)
        we3 = wb.create_sheet('sheet3', 2)

        if choice == '单位':
            
            we3.title = '水利工程运行维护'
            we2.title = '占比'
            we.title = '数量'
            we['A1'] = '全部诉求类型'
            we2['A1'] = '全部诉求类型'
            we3['A1'] = '道路破损'
            we['A2'] = '单位'
            we2['A2'] = '单位'
            we3['A2'] = '单位'
        elif choice == '诉求':
            we2.title = '占比'
            we.title = '数量'
            we['A1'] = '全部诉求'
            we['A2'] = '诉求类型（一级诉求）'
            we2['A1'] = '全部诉求'
            we2['A2'] = '诉求类型（一级诉求）'
        month_num = excel_num
        for i in range(0, month_num):
            we.cell(2, 2 + i).value = filelist[i]
            we2.cell(2, 2 + i).value = filelist[i]
            we3.cell(2, 2 + i).value = filelist[i]
            i += 1
        for j in range(0, len(getdict(dict_filepath))):
            we.cell(3 + j, 1).value = getdict(dict_filepath)[j]['名称']
            we2.cell(3 + j, 1).value = getdict(dict_filepath)[j]['名称']
            we3.cell(3 + j, 1).value = getdict(dict_filepath)[j]['名称']
            j += 1
        first_num = j + 4
        if choice == '单位':
            we.cell(first_num, 1).value = '非权属类型'  # 非权属部分   相邻部分空一格
            we.cell(first_num + 1, 1).value = '单位'
            we2.cell(first_num, 1).value = '非权属类型'  # 非权属部分   相邻部分空一格
            we2.cell(first_num + 1, 1).value = '单位'
            we3.cell(first_num, 1).value = '路灯'  # 路灯
            we3.cell(first_num + 1, 1).value = '单位'
        elif choice == '诉求':
            we.cell(first_num, 1).value = '非权属诉求'  # 非权属部分   相邻部分空一格
            we.cell(first_num + 1, 1).value = '诉求类型（一级诉求）'
            we2.cell(first_num, 1).value = '非权属诉求'  # 非权属部分   相邻部分空一格
            we2.cell(first_num + 1, 1).value = '诉求类型（一级诉求）'
        for i in range(0, month_num):
            we.cell(first_num + 1, 2 + i).value = filelist[i]
            we2.cell(first_num + 1, 2 + i).value = filelist[i]
            we3.cell(first_num + 1, 2 + i).value = filelist[i]
            i += 1
        for j in range(0, len(getdict(dict_filepath))):
            we.cell(2 + first_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            we2.cell(2 + first_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            we3.cell(2 + first_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            j += 1
        second_num = first_num + j + 3  # 已解决+剔除+挂账 部分
        if choice == '单位':
            we.cell(second_num, 1).value = '已解决+剔除+挂账'  # 非权属部分
            we.cell(second_num + 1, 1).value = '单位'
            we2.cell(second_num, 1).value = '已解决+剔除+挂账'  # 非权属部分
            we2.cell(second_num + 1, 1).value = '单位'
            we3.cell(second_num, 1).value = '路边乱停车'  # 非权属部分
            we3.cell(second_num + 1, 1).value = '单位'
        elif choice == '诉求':
            we.cell(second_num, 1).value = '已解决+剔除+挂账'  # 非权属部分
            we.cell(second_num + 1, 1).value = '诉求类型（一级诉求）'
            we2.cell(second_num, 1).value = '已解决+剔除+挂账'  # 非权属部分
            we2.cell(second_num + 1, 1).value = '诉求类型（一级诉求）'
        for i in range(0, month_num):
            we.cell(second_num + 1, 2 + i).value = filelist[i]
            we2.cell(second_num + 1, 2 + i).value =filelist[i]
            we3.cell(second_num + 1, 2 + i).value = filelist[i]
            i += 1
        for j in range(0, len(getdict(dict_filepath))):
            we.cell(2 + second_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            we2.cell(2 + second_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            we3.cell(2 + second_num + j, 1).value = getdict(dict_filepath)[j]['名称']
            j += 1

        if choice == '单位':
            third_num = second_num + j + 3
            we.cell(third_num, 1).value = '水环境维护类 总诉求'  # 水环境维护类 总诉求
            we.cell(third_num + 1, 1).value = '单位'
            we2.cell(third_num, 1).value = '水环境维护类 总诉求'  # 水环境维护类 总诉求
            we2.cell(third_num + 1, 1).value = '单位'
            for i in range(0, month_num):
                we.cell(third_num + 1, 2 + i).value = filelist[i]
                we2.cell(third_num + 1, 2 + i).value = filelist[i]
                i += 1
            for j in range(0, len(getdict(dict_filepath))):
                we.cell(2 + third_num + j, 1).value = getdict(dict_filepath)[j]['名称']
                we2.cell(2 + third_num + j, 1).value = getdict(dict_filepath)[j]['名称']
                j += 1
            fourth_num = third_num + j + 3
            we.cell(fourth_num, 1).value = '水环境维护类 已解决+剔除+挂账'  # 水环境维护类 已解决+剔除+挂账
            we.cell(fourth_num + 1, 1).value = '单位'
            we2.cell(fourth_num, 1).value = '水环境维护类 已解决+剔除+挂账'  # 水环境维护类 已解决+剔除+挂账
            we2.cell(fourth_num + 1, 1).value = '单位'
            for i in range(0, month_num):
                we.cell(fourth_num + 1, 2 + i).value = filelist[i].replace('.xlsx', '')
                we2.cell(fourth_num + 1, 2 + i).value = filelist[i].replace('.xlsx', '')
                i += 1
            for j in range(0, len(getdict(dict_filepath))):
                we.cell(2 + fourth_num + j, 1).value = getdict(dict_filepath)[j]['名称']
                we2.cell(2 + fourth_num + j, 1).value = getdict(dict_filepath)[j]['名称']
                j += 1
        wb.save(excel_filepath)  # 保存表格
        return (first_num)

    def getdict(filepath):
        book = load_workbook(filepath)
        sheet = book.active
        rows = sheet.rows
        headers = [cell.value for cell in next(rows)]
        all_rows = []
        for row in rows:
            # 创建一个空字典来存放这里面的值
            data = {}
            for title, cell in zip(headers, row):
                data[title] = cell.value
            all_rows.append(data)
        # print(len(all_rows),all_rows[62]['名称'])
        # print(all_rows )
        return (all_rows)

    def town_getdict(type):
        book = openpyxl.load_workbook(f"./demo_excel/街道三级诉求/{type}三级诉求.xlsx")
        # sheet = book.active
        sheet = book['Sheet1']
        rows = sheet.rows
        headers = [cell.value for cell in next(rows)]
        all_rows = []
        for row in rows:
            # 创建一个空字典来存放这里面的值
            data = {}
            for title, cell in zip(headers, row):
                data[title] = cell.value
            all_rows.append(data)
        # print(len(all_rows),all_rows[62]['名称'])
        # print(all_rows   )
        return (all_rows)

    def int_str(int_num):
        lst = []
        s = ''
        if int_num == 0:
            s = '0'
        while int_num > 0:
            num = int_num % 10
            lst.append(num)
            int_num //= 10
        for i in lst[::-1]:
            s += str(i)
        # print(s)
        return s

    def percent(num1, num2):
        a = '{:.2%}'.format(num1 / num2)
        # print(a)
        return (a)
    # monthlist=['']
    filelist = ['2020.1','2020.2','2020.3','2020.4','2020.5','2020.6','2020.7','2020.8','2020.9','2020.10','2020.11','2020.12','2021.1','2021.2','2021.3','2021.4','2021.5','2021.6','2021.7','2021.8','2021.9','2021.10','2021.11','2021.12','2022.1','2022.2','2022.3','2022.4','2022.5','2022.6','2022.7','2022.8','2022.9','2022.10','2022.11','2022.12','2023.1','2023.2','2023.3','2023.4','2023.5','2023.6','2023.7','2023.8','2023.9','2023.10','2023.11','2023.12']
    if year==2020:
        filelist=filelist[:month]
    elif year==2021:
        filelist=filelist[:12+month]
    elif year==2022:
        filelist=filelist[:24+month]
    elif year==2023:
        filelist=filelist[:36+month]
    # for root, dirs, files in os.walk("./data", topdown=False):
    # for root, dirs, files in os.walk(InDirname, topdown=False):
    #     for name in files:
    #         str2 = os.path.join(name)  # str = os.path.join(root, name)

    #         if str2.split('.')[-1] == 'xlsx':
    #             filelist.append(str2)
    excel_num = len(filelist)

    # print(filelist, excel_num)  # 这里filelist从0开始，到（excel数量-1）结束

    def alltongji(choice, first_path, second_path, third_path):  #### 程序运行时不要打开用到的 Excel ！
        rb = openpyxl.load_workbook(first_path)
        if choice == '水利工程运行维护' or choice=='单位':
            wl = rb['数量']
            wl2 = rb['占比']
            wl3 = rb['水利工程运行维护']
        if choice == '诉求':
            wl = rb['数量']
            wl2 = rb['占比']
        for i in range(excel_num - 1, excel_num):  # 决定统计表格起始    *********************************************
            if choice == '单位':
                new_num = getexcel('单位', first_path, second_path)
            elif choice == '诉求':
                new_num = getexcel('诉求', first_path, second_path)
            elif choice == '水利工程运行维护':
                new_num = getexcel('单位', first_path, second_path)  # 因为同样是统计个办公单位的
            factor_column = 0  ##  '主办单位'所在列
            result_column = 0  ##
            data = openpyxl.load_workbook(InDirname)
            sheets = data.sheetnames
            sheet = data[sheets[0]]
            # sheet = data["sheet1"]
            column = len(sheet[1])  # 表格列数
            row = len(tuple(sheet))  # 表格行数
            for column_num in range(1, column + 1):
                if sheet.cell(row=1, column=column_num).value == '主办单位':
                    # print(sheet[1][column_num-1].value)
                    factor_column = column_num
            # print(sheet.cell(row=12,column=factor_column).value)     #测试空格单元 类型是 ‘ ‘ 还是None，none会出错。
            for column_num in range(1, column + 1):
                if sheet.cell(row=1, column=column_num).value == '解决类型':
                    # print(sheet[1][column_num-1].value)
                    result_column = column_num
            # print(sheet.cell(row=41, column=result_column).value)
            for column_num in range(1, column + 1):
                if sheet.cell(row=1, column=column_num).value == '诉求类型一级':
                    # print(sheet[1][column_num-1].value)
                    type_column = column_num
            for column_num in range(1, column + 1):
                if sheet.cell(row=1, column=column_num).value == '诉求类型三级':
                    third_type_column = column_num
            # print( factor_column, result_column,type_column,row,column)
            factor_len = len(getdict(second_path))  # 办理单位数量
            # print(factor_len)
            print(f"{choice}{filelist[i]}开始统计···")
            all_tongji, not_tongji, other_tongji, type_all_tongji, type_other_tongji = 0, 0, 0, 0, 0
            all_list = {}
            not_list = {}
            other_list = {}
            type_all_list = {}
            type_other_list = {}  # 存储每月各单位的各类型数量
            kong_num = 0  # 办公单位为空
            for j in range(0, factor_len):  # 循环所有办公单位
                all_num = 0  # 全部类型
                not_num = 0  # 非权属
                other_num = 0  # 已解决+剔除+挂账 （等）
                type_all_num, type_other_num, type_not_num, kong_num1, kong_num2 = 0, 0, 0, 0, 0
                all_data = []
                not_data = []
                other_data = []
                type_all_data = []
                type_other_data = []
                for k in range(2, row + 1):
                    if choice == '单位':
                        if format(sheet.cell(row=k, column=factor_column).value.replace('（', '').replace('）',
                                                                                                         '')) == format(
                                getdict(second_path)[j]['名称'].replace('(', '').replace(')',
                                                                                       '')):  # 办理单位所在字典  一个去中文（） ,一个去英文()
                            all_num += 1
                            if format(sheet.cell(row=k, column=result_column).value.replace('（', '').replace('）',
                                                                                                             '')) == '非权属':
                                not_num += 1
                            elif format(sheet.cell(row=k, column=result_column).value.replace('（', '').replace('）',
                                                                                                               '')) == '已解决' or '剔除' or '挂账' or '' or '主动剔除' or 'none' or '正在办理':
                                other_num += 1
                            if sheet.cell(row=k, column=type_column).value == '水环境维护':
                                type_all_num += 1
                                if format(sheet.cell(row=k, column=result_column).value.replace('（', '').replace('）',
                                                                                                                 '')) == '非权属':
                                    type_not_num += 1
                        elif format(
                                sheet.cell(row=k, column=factor_column).value.replace('（', '').replace('）', '')) == '':
                            kong_num1 += 1
                        elif sheet.cell(row=k, column=factor_column).value == 'none':
                            kong_num2 += 1
                        type_other_num = type_all_num - type_not_num

                    elif choice == '诉求':
                        if format(sheet.cell(row=k, column=type_column).value.replace('（', '').replace('）',
                                                                                                       '')) == format(
                                getdict(second_path)[j]['名称'].replace('(', '').replace(')',
                                                                                       '')):  # 办理单位所在字典  一个去中文（） ,一个去英文()
                            all_num += 1
                            if format(sheet.cell(row=k, column=result_column).value.replace('（', '').replace('）',
                                                                                                             '')) == '非权属':
                                not_num += 1
                            elif format(sheet.cell(row=k, column=result_column).value.replace('（', '').replace('）',
                                                                                                               '')) == '已解决' or '剔除' or '挂账' or '' or '主动剔除' or 'none' or '正在办理':
                                other_num += 1
                        elif format(
                                sheet.cell(row=k, column=type_column).value.replace('（', '').replace('）', '')) == '':
                            kong_num1 += 1
                        elif sheet.cell(row=k, column=type_column).value == 'none':
                            kong_num2 += 1
                        type_other_num = type_all_num - type_not_num

                    elif choice == '水利工程运行维护':
                        if format(sheet.cell(row=k, column=factor_column).value.replace('（', '').replace('）',
                                                                                                         '')) == format(
                                getdict(second_path)[j]['名称'].replace('(', '').replace(')',
                                                                                       '')):  # 办理单位所在字典  一个去中文（） ,一个去英文()
                            if sheet.cell(row=k, column=type_column).value == '水利工程运行维护':
                                if sheet.cell(row=k, column=third_type_column).value == '道路破损':
                                    all_num += 1
                                if sheet.cell(row=k, column=third_type_column).value == '路灯':
                                    not_num += 1
                                if sheet.cell(row=k, column=third_type_column).value == '路边乱停车':
                                    other_num += 1
                        elif format(
                                sheet.cell(row=k, column=factor_column).value.replace('（', '').replace('）', '')) == '':
                            kong_num1 += 1
                        elif sheet.cell(row=k, column=factor_column).value == 'none':
                            kong_num2 += 1
                        type_other_num = type_all_num - type_not_num

                ##  结果存入数组
                all_list[j] = all_num
                all_data.append(all_list)
                not_list[j] = not_num
                not_data.append(not_list)
                other_list[j] = other_num
                other_data.append(other_list)
                type_all_list[j] = type_all_num
                type_all_data.append(type_all_list)
                type_other_list[j] = type_other_num
                type_other_data.append(type_other_list)
                all_tongji += all_num
                not_tongji += not_num
                other_tongji += other_num
                type_all_tongji += type_all_num
                type_other_tongji += type_other_num
                kong_num = kong_num1 + kong_num2
                print(all_num, not_num, other_num, all_tongji, not_tongji, other_tongji, kong_num, type_all_num,
                      type_other_num, row - 1, j)
                if type_all_tongji == 0:
                    type_all_tongji = 1
                if type_other_tongji == 0:
                    type_other_tongji = 1

            for j in range(0, factor_len):
                if choice == '单位':
                    month_num = excel_num
                    for ii in range(0, month_num):
                        wl.cell(2, 2 + ii).value = filelist[ii]
                        wl2.cell(2, 2 + ii).value = filelist[ii]
                        wl3.cell(2, 2 + ii).value = filelist[ii]

                        wl.cell(new_num + 1, 2 + ii).value = filelist[ii]
                        wl2.cell(new_num + 1, 2 + ii).value = filelist[ii]
                        wl3.cell(new_num + 1, 2 + ii).value = filelist[ii]

                        wl.cell(2 * new_num  , 2 + ii).value = filelist[ii]
                        wl2.cell(2 * new_num , 2 + ii).value = filelist[ii]
                        wl3.cell(2 * new_num , 2 + ii).value = filelist[ii]

                        ######单位独有
                        wl.cell(3 * new_num- 1, 2 + ii).value = filelist[ii]
                        wl2.cell(3 * new_num- 1, 2 + ii).value = filelist[ii]

                        wl.cell(4 * new_num - 2, 2 + ii).value = filelist[ii]
                        wl2.cell(4 * new_num - 2, 2 + ii).value = filelist[ii]
                    #   数量
                    wl.cell(3 + j, 2 + i).value = int_str(all_data[0][j])  # 统计结果写入 汇总表
                    wl.cell(new_num + 2 + j, 2 + i).value = int_str(not_data[0][j])
                    wl.cell(2 * new_num + 1 + j, 2 + i).value = int_str(other_data[0][j])
                    #   占比
                    wl2.cell(3 + j, 2 + i).value = percent(all_data[0][j], all_tongji)  # 统计结果写入 汇总表
                    wl2.cell(new_num + 2 + j, 2 + i).value = percent(not_data[0][j], not_tongji)
                    wl2.cell(2 * new_num + 1 + j, 2 + i).value = percent(other_data[0][j], other_tongji)
                    #   水环境维护
                    wl.cell(3 * new_num + j, 2 + i).value = int_str(type_all_data[0][j])
                    wl.cell(4 * new_num - 1 + j, 2 + i).value = int_str(type_other_data[0][j])
                    wl2.cell(3 * new_num + j, 2 + i).value = percent(type_all_data[0][j], type_all_tongji)
                    wl2.cell(4 * new_num - 1 + j, 2 + i).value = percent(type_other_data[0][j], type_other_tongji)

                elif choice == '诉求':
                    month_num = excel_num
                    for ii in range(0, month_num):
                        wl.cell(2, 2 + ii).value = filelist[ii]
                        wl2.cell(2, 2 + ii).value = filelist[ii]


                        wl.cell(new_num + 1, 2 + ii).value = filelist[ii]
                        wl2.cell(new_num + 1, 2 + ii).value = filelist[ii]


                        wl.cell(2 * new_num, 2 + ii).value = filelist[ii]
                        wl2.cell(2 * new_num, 2 + ii).value = filelist[ii]


                    #   数量
                    wl.cell(3 + j, 2 + i).value = int_str(all_data[0][j])  # 统计结果写入 汇总表
                    wl.cell(new_num + 2 + j, 2 + i).value = int_str(not_data[0][j])
                    wl.cell(2 * new_num + 1 + j, 2 + i).value = int_str(other_data[0][j])
                    #   占比
                    wl2.cell(3 + j, 2 + i).value = percent(all_data[0][j], all_tongji)  # 统计结果写入 汇总表
                    wl2.cell(new_num + 2 + j, 2 + i).value = percent(not_data[0][j], not_tongji)
                    wl2.cell(2 * new_num + 1 + j, 2 + i).value = percent(other_data[0][j], other_tongji)

                elif choice == '水利工程运行维护':
                    month_num = excel_num
                    for ii in range(0, month_num):
                        wl.cell(2, 2 + ii).value = filelist[ii]
                        wl2.cell(2, 2 + ii).value = filelist[ii]
                        wl3.cell(2, 2 + ii).value = filelist[ii]

                        wl.cell(new_num + 1, 2 + ii).value = filelist[ii]
                        wl2.cell(new_num + 1, 2 + ii).value = filelist[ii]
                        wl3.cell(new_num + 1, 2 + ii).value = filelist[ii]

                        wl.cell(2 * new_num, 2 + ii).value = filelist[ii]
                        wl2.cell(2 * new_num, 2 + ii).value = filelist[ii]
                        wl3.cell(2 * new_num, 2 + ii).value = filelist[ii]

                        ######单位独有
                        wl.cell(3 * new_num - 1, 2 + ii).value = filelist[ii]
                        wl2.cell(3 * new_num - 1, 2 + ii).value = filelist[ii]

                        wl.cell(4 * new_num - 2, 2 + ii).value = filelist[ii]
                        wl2.cell(4 * new_num - 2, 2 + ii).value = filelist[ii]
                    #   数量
                    wl3.cell(3 + j, 2 + i).value = int_str(all_data[0][j])  # 统计结果写入 汇总表
                    wl3.cell(new_num + 2 + j, 2 + i).value = int_str(not_data[0][j])
                    wl3.cell(2 * new_num + 1 + j, 2 + i).value = int_str(other_data[0][j])
            #   保存
            rb.save(third_path)
            if choice == '单位':
                rb.save(filepath3)
            if choice == '诉求':
                rb.save(filepath4)
            if choice == '水利工程运行维护':
                rb.save(filepath3)

            if (all_tongji + kong_num) != row - 1:
                if choice == '单位':
                    print(
                        f"此月份表格有{row - 1 - all_tongji - kong_num}条办公单位信息未添加到‘办公单位.xlsx’文件中，或有的单位名字打错。")  # 这里格式要求严格，如果名字没错也没有空的，那么注意名字前后的 ‘ ’，即，空格！
                elif choice == '诉求':
                    print(f"此月份表格有{row - 1 - all_tongji - kong_num}条一级诉求类型信息未添加到‘诉求类型一级.xlsx’文件中，或有的诉求类型名字打错。")
                elif choice == '水利工程运行维护':
                    print(f"水利工程运行维护--此月份表格有{all_tongji}条关于道路破损，{not_tongji}条关于道路破损，{other_tongji}条关于道路破损。")
            if kong_num != 0:
                if choice == '单位':
                    print(f"{filelist[i]}表中办公单位为空的数据有{kong_num}条,包括’ ‘和’none‘！")
                elif choice == '诉求':
                    print(f"{filelist[i]}表中一级诉求类型为空的数据有{kong_num}条,包括’ ‘和’none‘！")
                elif choice == '水利工程运行维护':
                    print(f"水利工程运行维护--{filelist[i]}表中办公单位为空的数据有{kong_num}条,包括’ ‘和’none‘！")
            print(f"{filelist[i]}统计完成")
        # rb.save(last_filepath)
        print('统计完成！')

    # if __name__ == '__main__':
    if  count_type=='单位解决类型':
        #  用于统计各单位全部诉求类型、非全属、解决等数量和占比，还有水环境维护类的全部诉求类型、非全属、解决等数量和占比。
        alltongji('单位', filepath3, filepath1, filepath8)  # 汇总样板表、办公单位、统计报告
    if count_type == '诉求变化趋势':
        #  用于统计诉求类型（一级诉求）的全部诉求类型、非全属、解决等数量和占比
        alltongji('诉求', filepath4, filepath2, filepath9)  # 诉求类型样板、诉求类型一级表、诉求风险趋势变化表
    if count_type == '水利工程运行维护':
        #  用于统计水利工程运行维护里各单位对道路破损、路灯、路边乱停车的数量统计
        alltongji('水利工程运行维护', filepath3, filepath1, filepath8)  # 汇总样板表、办公单位、统计报告
    if count_type == '各区街道集团':
        getexcel_town('各区街道集团', filepath6, filepath5)  # 样板
        #  用于统计各区街道集团的二级问题诉求，同时对统计的二级问题进行排序，然后对排序后的二级问题对应的三级问题进行统计
        alltongji_town('各区街道集团', 1, filepath6, filepath5, filepath7)  # 标题、样表、起始月份、字典、输出表

def Draw(data_type='直派水务局',draw_type='折线图（诉求变化趋势）',Infilename=None,Outdirname=None):
    # import matplotlib.pyplot as plt
    plt.rcParams['font.sans-serif'] = ['SimHei']
    import os
    if not os.path.exists(Outdirname+'\\绘图\\'):
        os.makedirs(Outdirname+'\\绘图\\')
    def scrdf(df,indexname,judge):#
        df0=df[indexname]
        indexs = []
        for index, value in enumerate(df0):
            if judge == value:
                indexs.append(index)
        scrdf = pd.DataFrame(df.iloc[indexs].values)
        scrdf.columns=df.columns
        return scrdf
    def Draw_line(name):
        def draw_line_base(month,datadf,yearlist,length):
            # plt.figure()
            plt.plot()
            step = int(length/12)
            # print(length,step)
            for i in range(step):
                # print("%d",i*12)
                plt.plot(month[:12],datadf[int(i*12):int((i+1)*12)], marker='o', markersize=3)
                for a, b in zip(month[:12], datadf[int(i*12):int((i+1)*12)]):
                    plt.text(a, b, b, ha='center', va='bottom', fontsize=10)  # 设置数据标签位置及大小
            plt.plot(month[:length-step*12],datadf[step*12:length],marker='o',markersize=3)
            for a, b in zip(month[:length-step*12], datadf[step*12:length]):
                plt.text(a, b, b, ha='center', va='bottom', fontsize=10)
            plt.legend(yearlist[:step+1])  # 设置折线名称
            # plt.show()
        df=pd.read_excel('./data/诉求风险趋势变化表.xlsx')[-32:]
        plt.rcParams['font.sans-serif'] = ['SimHei']
        ptdf = df.iloc[df['全部诉求'].values.tolist().index(name)]
        datadf = ptdf.values.tolist()
        datadf.remove(name)
        datadf = list(map(int, datadf))
        yearlist=['2020年','2021年','2022年','2023年','2024年','2025年','2026年','2027年','2028年']
        month = ['1月', '2月', '3月', '4月', '5月', '6月', '7月', '8月', '9月', '10月', '11月', '12月']
        length=len(datadf)
        draw_line_base(month,datadf,yearlist,length)
        plt.title(label=name)
        plt.xlabel('时间')  # x轴标题
        plt.ylabel('诉求值')  # y轴标题
        plt.savefig(Outdirname  + '\\绘图\\'+name+'诉求变化趋势图.jpg')
    def Draw_pie(data_type='直派水务局'): # make a square figure
        df=pd.read_excel(Infilename)
        if data_type=='直派水务局':
            pt0=df['诉求类型一级'].value_counts()
            pt1=df['诉求类型二级'].value_counts()
            pt2=df['诉求类型三级'].value_counts()
        elif data_type=='直派行业':
            pt0=df['一级问题'].value_counts()
            pt1=df['二级问题'].value_counts()
            pt2=df['三级问题'].value_counts()
        # For China, make the piece explode a bit
        expl = [0.1,0.05,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]#第二块即China离开圆心0.1
        # Colors used. Recycle if not enough.
        colors  = ["firebrick","indianred","tomato","rosybrown","lightcoral","salmon","sienna","sandybrown",'peru','greenyellow','palegreen','springgreen','aquamarine','cyan','deepskyblue',"lightskyblue",'violet','hotpink',"coral","magenta","green","lightgreen","orange"]  #设置颜色（循环显示）
        # Pie Plot
        # autopct: format of "percent" string;百分数格式
        lb0=pt0.index.tolist()
        qs0=pt0.values.tolist()
        lb1=pt1.index.tolist()
        qs1=pt1.values.tolist()
        lb2=pt2.index.tolist()
        qs2=pt2.values.tolist()
        if len(lb0)>=15:
            qs00=[]
            qs00.extend(qs0[:15])
            qs00.append(sum(qs0[15:]))
            lb00=[]
            lb00.extend(lb0[:15])
            lb00.append('其余诉求')
            plt.figure(1, figsize=(9,6))
            plt.pie(qs00,colors= colors,explode=expl[:len(qs00)], labels=lb00, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局一级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right",bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局一级诉求-类型占比分布图.png')
            plt.close()
        else:
            plt.figure(1, figsize=(9,6))
            plt.pie(qs0,colors= colors,explode=expl[:len(qs0)], labels=lb0, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局一级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right",bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局一级诉求-类型占比分布图.png')
            plt.close()
        if len(lb1)>=20:
            qs01=[]
            qs01.extend(qs1[:20])
            qs01.append(sum(qs1[20:]))
            lb01=[]
            lb01.extend(lb1[:20])
            lb01.append('其余诉求')
            plt.figure(1, figsize=(9,6))
            plt.pie(qs01,colors= colors,explode=expl[:len(qs01)], labels=lb01, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局二级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right", 
                            bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局二级诉求-类型占比分布图.png')
            plt.close()
        else:
            plt.figure(1, figsize=(9,6))
            plt.pie(qs1,colors= colors,explode=expl[:len(qs1)], labels=lb1, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局二级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right",bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局二级诉求-类型占比分布图.png')
            plt.close()
        if len(lb2)>=20:
            qs02=[]
            qs02.extend(qs2[:20])
            qs02.append(sum(qs2[20:]))
            lb02=[]
            lb02.extend(lb2[:20])
            lb02.append('其余诉求')
            plt.figure(1, figsize=(9,6))
            plt.pie(qs02,colors= colors,explode=expl[:len(qs02)], labels=lb02, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局三级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right", 
                            bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局三级诉求-类型占比分布图.png')
            plt.close()
        else:
            plt.figure(1, figsize=(9,6))
            plt.pie(qs2,colors= colors,explode=expl[:len(qs2)], labels=lb2, autopct='%1.1f%%',pctdistance=1.1,labeldistance=1.2, shadow=True)
            plt.title('直派水务局三级诉求-类型占比分布图',y=-0.1, bbox={'facecolor':'0.8', 'pad':5})
            plt.legend(bbox_to_anchor=(1,1), loc="upper right",bbox_transform=plt.gcf().transFigure,fontsize=10,borderaxespad=0.1)
            plt.savefig(Outdirname+'\\绘图\\直派水务局三级诉求-类型占比分布图.png')
            plt.close()
    def Draw_bar(data_type='直派水务局'):
        
        def draw_bar(pt,name):
            plt.rcParams['font.sans-serif'] = ['SimHei']
            # plt.figure(14.0,8.0)
            pt.plot.bar()
            x = range(0, len(pt.index.tolist()), 1)
            plt.xticks(x,rotation=60)
            plt.subplots_adjust(bottom=0.35)
            plt.title(name)
            for a, b in zip(x, pt.values.tolist()):
                    plt.text(a, b+0.05,'%d' %b, ha='center', va='bottom', fontsize=10)
            plt.savefig(Outdirname + '\\绘图\\'+name+'-诉求柱状图.jpg')
            plt.close()
        if data_type=='直派水务局':
            df=pd.read_excel(Infilename)
            pt = df['诉求类型一级'].value_counts()
            draw_bar(pt,'直派水务局诉求类型一级')
            pt1list=pt.index.tolist()
            for i in range(len(pt1list)):
                draw_bar(scrdf(df,'诉求类型一级',pt1list[i])['诉求类型二级'].value_counts(),'直派水务局诉求类型二级-'+pt1list[i])
        elif data_type=='直派行业':
            df=pd.read_excel(Infilename)
            pt = df['二级问题'].value_counts()
            draw_bar(pt,'直派行业二级问题')
            pt1list=pt.index.tolist()
            for i in range(len(pt1list)):
                draw_bar(scrdf(df,'二级问题',pt1list[i])['三级问题'].value_counts(),'直派行业三级问题-'+pt1list[i])
    def Draw_geo():
        from pyecharts import options as opts
        from pyecharts.charts import Geo, Map
        from pyecharts.render import make_snapshot
        from snapshot_selenium import snapshot as driver
        def draw_geo(df,name):
            quname=['密云区','延庆区','朝阳区','丰台区','石景山区','海淀区','门头沟区','房山区','通州区','顺义区','昌平区','大兴区','怀柔区','平谷区','东城区','西城区']
            ptindex=df['被反映区'].value_counts().index.tolist()
            # ptindex=list(set(ptindex+quname))
            ptindex=list((ptindex+ [item for item in quname if str(item) not in ptindex]))
            ptvalue=df['被反映区'].value_counts().values.tolist()
            if len(ptindex)!=len(ptvalue):
                ptvalue.extend(0 for _ in range(len(ptindex)-len(ptvalue)))
            maxvalue=max(map(int,ptvalue))
            while (maxvalue %5):
                maxvalue+=1
            c = (
                Map()
                .add("", [list(z) for z in zip(ptindex, ptvalue)], "北京",zoom=1.25,aspect_scale=0.9,layout_center=['60%','60%'],
                label_opts=opts.LabelOpts(is_show=True,position='outsideleft',font_size=10,color='#708069',rotate = '30',horizontal_align = 'center',font_weight='bold',vertical_align ='middle'))
                .set_global_opts(
                title_opts=opts.TitleOpts(title=name),
                visualmap_opts=opts.VisualMapOpts(type_='color', orient='vertical',max_=maxvalue,
                                                    is_piecewise=True,range_color=['#71ae9b','#8bb08e', '#c1d389','#e3d76b','#f7b44a','#f27127','#cc8775']
                ))
            )
            make_snapshot(driver, c.render(), Outdirname +'\\pic\\'+ name+".png")
            b = (
                Map()
                .add("", [list(z) for z in zip(ptindex, ptvalue)], "北京",zoom=1.25,aspect_scale=0.9,layout_center=['60%','60%'],
                label_opts=opts.LabelOpts(is_show=True,position='outsideleft',font_size=10,color='#708069',rotate = '30',horizontal_align = 'center',font_weight='bold',vertical_align ='middle'))
                .set_global_opts(
                title_opts=opts.TitleOpts(title=name),
                visualmap_opts=opts.VisualMapOpts(type_='color', orient='vertical',is_piecewise=True,pieces=[{'min':0,'max':10,'label':'[0-10)',"color":'#71ae9b'},{'min':11,'max':30,'label':'[10-30)',"color":'#8bb08e'},{'min':31,'max':100,'label':'[30-100)',"color":'#c1d389'},{'min':101,'max':200,'label':'[100-200)',"color":'#e3d76b'},{'min':201,'max':350,'label':'[200-350)',"color":'#f7b44a'},
                                                            {'min':350,'label':'[350,)',"color":'#f27127'}]
                                                            # {'min':0,'max'=10,label:'0-10','#cc8775']
                                                    )
            ))
            make_snapshot(driver, b.render(), Outdirname +'\\pic\\'+ name +".png")
        df=pd.read_excel(Infilename)
        pt = df['三级问题'].value_counts()
        pt1list=pt.index.tolist()
        for i in range(10):
            draw_geo(scrdf(df,'三级问题',pt1list[i]),'直派行业三级诉求-'+pt1list[i]+'-地理分布图')# if __name__=='__main__':
    if draw_type=='水利工程运行维护变化趋势-折线图':
        Draw_line('水利工程运行维护')
    elif draw_type=='水环境维护变化趋势-折线图':
        Draw_line('水环境维护')
    elif draw_type=='雨水与海绵城市变化趋势-折线图':
        Draw_line('雨水与海绵城市')
    elif draw_type=='饼状图':
        Draw_pie(data_type)
    elif draw_type=='柱状图（直派水务局一级诉求）' or draw_type=='柱状图（直派行业三级问题）'  :
        Draw_bar(data_type)
    elif draw_type=='地理分布图（区）' and data_type=='直派行业':
            Draw_geo()